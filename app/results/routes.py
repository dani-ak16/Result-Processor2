from flask import Blueprint, render_template, request, redirect, url_for
from PIL import Image, ExifTags
import zipfile, tempfile
from playwright.sync_api import sync_playwright
from weasyprint import HTML
import io
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from extensions import db
from flask import render_template, request, redirect, url_for
from . import results_bp

@results_bp.route('/generate-reports', methods=['GET', 'POST'])
def generate_reports():
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        class_arm_id = request.form.get('class_arm_id')
        term = int(request.form['term'])
        session = request.form['session']
        report_type = request.form.get('report_type', 'full_term')
        
        cursor.execute("SELECT class_id FROM class_arms WHERE id = ?", (class_arm_id,))
        class_result = cursor.fetchone()
        class_id = class_result['class_id']

        cursor.execute("SELECT level FROM classes WHERE id = ?", (class_id,))
        level = cursor.fetchone()
        class_level = level['level']

        results = []

        results = cursor.execute("""
                SELECT s.id, s.full_name, s.reg_number, c.name || ' ' || a.arm AS class_name,
                    AVG(sc.total_score) AS average
                FROM students s
                JOIN scores sc ON s.id = sc.student_id
                JOIN student_classes x ON s.id = x.student_id
                JOIN class_arms a ON x.class_arm_id = a.id
                JOIN classes c ON a.class_id = c.id
                WHERE a.class_id = ? AND x.term = ? AND x.session = ? AND sc.report_type = ?
                GROUP BY s.id
                ORDER BY average DESC
            """, (class_id, term, session, report_type)).fetchall()

        class_avg = sum(r["average"] for r in results)/len(results) if results else 0

        # Convert to list of tuples (student_id, average)
        entries = [(row['id'], row['average']) for row in results]

        # Sort by highest average
        entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)

        # Assign positions (1st, 2nd, 3rd…)
        rankings = {}
        position = 1

        for i, (student_id, avg) in enumerate(entries_sorted):
            # Handle ties: if same score as previous student, same rank
            if i > 0 and avg == entries_sorted[i - 1][1]:
                rankings[student_id] = rankings[entries_sorted[i - 1][0]]
            else:
                rankings[student_id] = position
            
            position += 1

        # --- Single Student Report ---
        if full_name:
            student = get_best_student_match(class_arm_id, session, full_name)
            print(student)

            if not student:
                return render_template("error.html", message="No student found")

            cursor.execute("""
                    SELECT sub.name AS subject, 
                           sc.ca1_score, sc.ca2_score, sc.ca3_score, sc.ca4_score, 
                           sc.exam_score, sc.total_score, sc.report_type
                    FROM scores sc
                    JOIN subjects sub ON sc.subject_id = sub.id
                    WHERE sc.student_id = ? AND sc.term = ? AND sc.session = ? AND sc.report_type = ?
                """, (student["id"], term, session, report_type))
            scores = cursor.fetchall()

            if not scores:
                return render_template("error.html", message="No results found")

            cursor.execute("""
                    SELECT days_present, days_absent, days_late, total_school_days
                    FROM attendance_summary 
                    WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
                """, (student['id'], class_arm_id, term, session))
            attendance_summary = cursor.fetchone()

            cursor.execute("""
                SELECT handwriting, sports_participation, practical_skills,
                    punctuality, politeness, neatness,
                    class_teacher_comment, principal_comment
                FROM student_assessments
                WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
            """, (student["id"], class_arm_id, term, session))
            assessment = cursor.fetchone()

            cursor.execute('''SELECT sk.name, ss.score
                FROM student_skills ss
                JOIN skills sk ON ss.skill_id = sk.id
                WHERE ss.student_id = ? AND ss.term = ? AND ss.session = ?''', (student["id"], term, session))
            skills = cursor.fetchone()

            total = sum(r["total_score"] for r in scores)
            average = total / len(scores)

            student_position = rankings.get(student['id'], None)

            if class_level == "JSS":
                position = student_position
                grade = None
            else:
                position = None
                grade = grade_from_average(average)


            # Choose template based on report type
            template_name = "half_term_report.html" if report_type == "half_term" else "full_term_report.html"
            logo_path = os.path.join(app.root_path, 'static', 'kembos_logo_nobg.png')


            return render_template(template_name,
                                           student=student,
                                           class_name=student["class_name"],
                                           scores=scores,
                                           term=term,
                                           logo_path=logo_path,
                                           session=session,
                                           average=average,
                                           class_average=class_avg,
                                           position=position,
                                           grade=grade,
                                           report_type=report_type,
                                           skills=skills,
                                           attendance_summary=attendance_summary,
                                           assessment=assessment,
                                           year=datetime.now().year,
                                           current_date=datetime.now().strftime("%Y-%m-%d"))

       # --- Whole Class Batch Reports ---
        cursor.execute("""
            SELECT s.id, s.reg_number, s.full_name, s.age, s.photo, 
                c.name || ' ' || a.arm AS class_name
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            JOIN class_arms a ON sc.class_arm_id = a.id
            JOIN classes c ON a.class_id = c.id
            WHERE sc.class_arm_id = ? AND sc.session = ?
            ORDER BY s.full_name
        """, (class_arm_id, session))
        students = cursor.fetchall()

        if not students:
            return render_template("error.html", message="No students found for this class/year")

        report_type = request.form.get("report_type", "full_term")

        # Sanitize ZIP filename
        safe_session = session.replace("/", "_")
        zip_filename = f"class_{class_arm_id}_term{term}_{safe_session}_reports.zip"

        # Prepare in-memory ZIP buffer
        zip_buffer = BytesIO()

        # Detect your current host (local or LAN)
        host_url = request.host_url.rstrip('/')

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for student in students:
                    reg_number = student["reg_number"]
                    
                    # Construct the exact URL for the student's report page
                    report_url = f"{host_url}/preview-report?student_id={student['id']}&term={term}&session={session}&report_type={report_type}"
                    
                    pdf_filename = f"{student['full_name'].replace(' ', '_')}_report.pdf"
                    pdf_path = os.path.join(tempfile.gettempdir(), pdf_filename)
                    
                    page = context.new_page()
                    print(f"🧾 Generating PDF for {student['full_name']} ({report_type})")
                    
                    try:
                        page.goto(report_url, wait_until="networkidle")
                        page.pdf(
                            path=pdf_path,
                            format="A4",
                            print_background=True,
                            margin={"top": "10mm", "bottom": "10mm", "left": "10mm", "right": "10mm"}
                        )
                        zf.write(pdf_path, pdf_filename)
                    except Exception as e:
                        print(f"❌ Error generating report for {student['full_name']}: {e}")
                    finally:
                        if os.path.exists(pdf_path):
                            os.remove(pdf_path)
                        page.close()
            
            browser.close()

        # Return ZIP as downloadable file
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype="application/zip"
        )


    # --- GET request → show report form ---
    cursor.execute("""
                SELECT a.id AS arm_id, c.name AS class_name, a.arm
                FROM class_arms a
                JOIN classes c ON a.class_id = c.id
                ORDER BY c.id, a.arm
            """)
    classes = cursor.fetchall()
    current_session=get_current_session()
    return render_template("report_form.html", classes=classes, current_session=current_session)

@results_bp.route("/preview-report")
def preview_report():
    student_id = request.args.get("student_id")
    term = request.args.get("term")
    session = request.args.get("session")
    report_type = request.args.get("report_type", "full_term")

    if not student_id or not term or not session:
        return "Missing parameters", 400

    db = get_db()
    cursor = db.cursor()

    # Get student info
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.age, s.photo, 
               c.name || ' ' || a.arm AS class_name, s.department_id, s.gender
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        JOIN class_arms a ON sc.class_arm_id = a.id
        JOIN classes c ON a.class_id = c.id
        WHERE s.id = ? AND sc.session = ?
    """, (student_id, session))
    student = cursor.fetchone()
    if not student:
        return "Student not found", 404

    # Get scores
    cursor.execute("""
        SELECT sub.name AS subject, 
               sc.ca1_score, sc.ca2_score, sc.ca3_score, sc.ca4_score,
               sc.exam_score, sc.total_score
        FROM scores sc
        JOIN subjects sub ON sc.subject_id = sub.id
        WHERE sc.student_id = ? AND sc.term = ? AND sc.session = ? AND sc.report_type = ?
    """, (student_id, term, session, report_type))
    scores = cursor.fetchall()

    # Get attendance
    cursor.execute("""
        SELECT days_present, days_absent, days_late, total_school_days
        FROM attendance_summary 
        WHERE student_id = ? AND class_arm_id = (
            SELECT class_arm_id FROM student_classes 
            WHERE student_id = ? AND session = ?
        )
        AND term = ? AND session = ?
    """, (student_id, student_id, session, term, session))
    attendance_summary = cursor.fetchone()

    cursor.execute("""
        SELECT handwriting, sports_participation, practical_skills,
            punctuality, politeness, neatness,
            class_teacher_comment, principal_comment
        FROM student_assessments
        WHERE student_id = ? AND term = ? AND session = ?
    """, (student["id"], term, session))
    assessment = cursor.fetchone()

    cursor.execute('''SELECT sk.name, ss.score
                FROM student_skills ss
                JOIN skills sk ON ss.skill_id = sk.id
                WHERE ss.student_id = ? AND ss.term = ? AND ss.session = ?''', (student["id"], term, session))
    skills = cursor.fetchone()
    
    cursor.execute("""
        SELECT class_arm_id
        FROM student_classes
        WHERE student_id = ?
    """, (student["id"],))
    class_arm = cursor.fetchone()

    cursor.execute("SELECT class_id FROM class_arms WHERE id = ?", (class_arm['class_arm_id'],))
    class_result = cursor.fetchone()
    class_id = class_result['class_id']

    cursor.execute("SELECT level FROM classes WHERE id = ?", (class_id,))
    level = cursor.fetchone()
    class_level = level['level']

    results = []

    results = cursor.execute("""
                SELECT s.id, s.full_name, s.reg_number, c.name || ' ' || a.arm AS class_name,
                    AVG(sc.total_score) AS average
                FROM students s
                JOIN scores sc ON s.id = sc.student_id
                JOIN student_classes x ON s.id = x.student_id
                JOIN class_arms a ON x.class_arm_id = a.id
                JOIN classes c ON a.class_id = c.id
                WHERE a.class_id = ? AND x.term = ? AND x.session = ? AND sc.report_type = ?
                GROUP BY s.id
                ORDER BY average DESC
            """, (class_id, term, session, report_type)).fetchall()

    # Convert to list of tuples (student_id, average)
    entries = [(row['id'], row['average']) for row in results]

    # Sort by highest average
    entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)

    # Assign positions (1st, 2nd, 3rd…)
    rankings = {}
    position = 1

    for i, (student_id, avg) in enumerate(entries_sorted):
        # Handle ties: if same score as previous student, same rank
        if i > 0 and avg == entries_sorted[i - 1][1]:
            rankings[student_id] = rankings[entries_sorted[i - 1][0]]
        else:
            rankings[student_id] = position
            
        position += 1

    class_avg = sum(r["average"] for r in results)/len(results) if results else 0
    total = sum(r["total_score"] or 0 for r in scores) if scores else 0
    average = total / len(scores) if scores else 0
    student_position = rankings.get(student["id"], None)

    if class_level == "JSS":
        position = student_position
        grade = None
    else:
        position = None
        grade = grade_from_average(average)

    template_name = "full_term_report.html" if report_type == "full_term" else "half_term_report.html"

    return render_template(template_name,
                           student=student,
                           class_name=student["class_name"],
                           scores=scores,
                           term=term,
                           session=session,
                           class_average=class_avg,
                           position=position,
                           grade=grade,
                           skills=skills,
                           average=average,
                           report_type=report_type,
                           attendance_summary=attendance_summary,
                           assessment=assessment,
                           year=datetime.now().year,
                           current_date=datetime.now().strftime("%Y-%m-%d"))

