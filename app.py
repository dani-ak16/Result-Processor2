import os
import socket
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, redirect, abort, url_for, g, session, flash, send_file, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
from playwright.sync_api import sync_playwright
import sqlite3
import base64
from PIL import Image, ExifTags
import zipfile, tempfile
import io
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from datetime import datetime
from weasyprint import HTML

app = Flask(__name__)

if os.environ.get('RENDER'):
    # Render production environment
    app.config['DATABASE'] = '/tmp/school_results.db' 
    app.config['UPLOAD_FOLDER'] = '/tmp/uploads/'  
    print("Running on RENDER environment")
else:
    # Local development environment
    app.config['DATABASE'] = 'school_results.db'
    app.config['UPLOAD_FOLDER'] = 'uploads/'
    print("Running on LOCAL environment")

app.config['ALLOWED_EXTENSIONS'] = {'csv', 'xlsx'}

# app.config['DATABASE'] = 'school_results copy.db'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'school_result_secret_key')

# Create necessary directories
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], "photos"), exist_ok=True)
    
# def get_db():
#     db = getattr(g, '_database', None)
#     if db is None:
#         db = g._database = sqlite3.connect(app.config['DATABASE'])
#         db.row_factory = sqlite3.Row
#     return db

def get_db():
    if 'db' not in g:
        # os.makedirs(os.path.dirname(app.config['DATABASE']), exist_ok=True)
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()

        #Users
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT UNIQUE NOT NULL,
                        password_hash TEXT NOT NULL,
                        role TEXT NOT NULL CHECK (
                            role IN ('admin', 'class_teacher', 'subject_teacher')
                        ),
                        is_active BOOLEAN DEFAULT 1,
                        created_at TEXT)''')

        # Classes table
        cursor.execute('''CREATE TABLE IF NOT EXISTS classes (
                        id INTEGER PRIMARY KEY,
                        name TEXT UNIQUE NOT NULL,
                        level TEXT NOT NULL)''')

        # Class arms table
        cursor.execute("""CREATE TABLE IF NOT EXISTS class_arms (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_id INTEGER NOT NULL,
                    arm TEXT NOT NULL,
                    UNIQUE(class_id, arm),
                    FOREIGN KEY (class_id) REFERENCES classes (id))""")
    
        # Students table
        cursor.execute("""CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    reg_number TEXT UNIQUE,
                    full_name TEXT NOT NULL,
                    age INTEGER,
                    gender TEXT,
                    photo TEXT,
                    department_id INTEGER,
                    FOREIGN KEY (department_id) REFERENCES departments (id))""")

        # Student → class assignment
        cursor.execute("""CREATE TABLE IF NOT EXISTS student_classes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id INTEGER NOT NULL,
                    class_arm_id INTEGER NOT NULL,
                    session TEXT NOT NULL,
                    term INTEGER NOT NULL,
                    UNIQUE(student_id, class_arm_id, session, term),
                    FOREIGN KEY (student_id) REFERENCES students (id),
                    FOREIGN KEY (class_arm_id) REFERENCES class_arms (id))""")

        # Subjects table
        cursor.execute("""CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            level TEXT NOT NULL CHECK(level IN ('junior', 'senior')),
            is_common_core BOOLEAN DEFAULT 0,
            UNIQUE(name, level))
        """)

        # Departments table
        cursor.execute("""CREATE TABLE IF NOT EXISTS departments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    level TEXT NOT NULL CHECK(level IN ('junior', 'senior')),
                    description TEXT)""")

        # Department-Subject relationships
        cursor.execute("""CREATE TABLE IF NOT EXISTS department_subjects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    department_id INTEGER NOT NULL,
                    subject_id INTEGER NOT NULL,
                    is_compulsory BOOLEAN DEFAULT 1,
                    UNIQUE(department_id, subject_id),
                    FOREIGN KEY (department_id) REFERENCES departments (id),
                    FOREIGN KEY (subject_id) REFERENCES subjects (id))""")

        # Class-Subject requirements
        cursor.execute("""CREATE TABLE IF NOT EXISTS class_subject_requirements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_arm_id INTEGER NOT NULL,
                    subject_id INTEGER NOT NULL,
                    is_compulsory BOOLEAN DEFAULT 1,
                    UNIQUE(class_arm_id, subject_id),
                    FOREIGN KEY (class_arm_id) REFERENCES class_arms (id),
                    FOREIGN KEY (subject_id) REFERENCES subjects (id))""")

        # Scores table
        cursor.execute('''CREATE TABLE IF NOT EXISTS scores (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER NOT NULL,
                        subject_id INTEGER NOT NULL,
                        class_arm_id INTEGER NOT NULL,
                        term INTEGER NOT NULL,
                        session TEXT NOT NULL,
                        ca1_score REAL NULL DEFAULT 0,
                        ca2_score REAL NULL DEFAULT 0,
                        ca3_score REAL NULL DEFAULT 0,
                        ca4_score REAL NULL DEFAULT 0,
                        exam_score REAL NULL DEFAULT 0,
                        total_score REAL DEFAULT 0,
                        report_type TEXT NOT NULL CHECK(report_type IN ('half_term', 'full_term')),
                        created_at TEXT NOT NULL,  
                        approved INTEGER DEFAULT 0,                     
                        UNIQUE(student_id, subject_id, class_arm_id, term, session, report_type),
                        FOREIGN KEY(student_id) REFERENCES students (id),
                        FOREIGN KEY(subject_id) REFERENCES subjects (id),
                        FOREIGN KEY(class_arm_id) REFERENCES class_arms (id))''')

        # Attendance summary table
        cursor.execute("""CREATE TABLE IF NOT EXISTS attendance_summary (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER NOT NULL,
                        class_arm_id INTEGER NOT NULL,
                        term INTEGER NOT NULL,
                        session TEXT NOT NULL,
                        days_present INTEGER DEFAULT 0,
                        days_absent INTEGER DEFAULT 0,
                        days_late INTEGER DEFAULT 0,
                        total_school_days INTEGER DEFAULT 0,
                        UNIQUE(student_id, class_arm_id, term, session),
                        FOREIGN KEY (student_id) REFERENCES students (id),
                        FOREIGN KEY (class_arm_id) REFERENCES class_arms (id))""")

        # Student assessments table
        cursor.execute("""CREATE TABLE IF NOT EXISTS student_assessments (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER NOT NULL,
                        class_arm_id INTEGER NOT NULL,
                        term INTEGER NOT NULL,
                        session TEXT NOT NULL,
                        handwriting INTEGER CHECK(handwriting BETWEEN 1 AND 5),
                        sports_participation INTEGER CHECK(sports_participation BETWEEN 1 AND 5),
                        practical_skills INTEGER CHECK(practical_skills BETWEEN 1 AND 5),
                        punctuality INTEGER CHECK(punctuality BETWEEN 1 AND 5),
                        politeness INTEGER CHECK(politeness BETWEEN 1 AND 5),
                        neatness INTEGER CHECK(neatness BETWEEN 1 AND 5),
                        class_teacher_comment TEXT,
                        principal_comment TEXT,
                        UNIQUE(student_id, class_arm_id, term, session),
                        FOREIGN KEY (student_id) REFERENCES students (id),
                        FOREIGN KEY (class_arm_id) REFERENCES class_arms (id))""")

        cursor.execute('''CREATE TABLE IF NOT EXISTS principal_comments (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            min_average REAL NOT NULL,
                            max_average REAL NOT NULL,
                            comment TEXT NOT NULL,
                            UNIQUE(min_average, max_average, comment))''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL)''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS student_skills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                class_arm_id INTEGER NOT NULL,
                term INTEGER NOT NULL,
                session TEXT NOT NULL,
                skill_id INTEGER NOT NULL,
                score INTEGER DEFAULT 0,
                UNIQUE(student_id, skill_id, class_arm_id, term, session),
                FOREIGN KEY (student_id) REFERENCES students(id),
                FOREIGN KEY (skill_id) REFERENCES skills(id)
            )''')

        skills = [
            "Coding",
            "Photography",
            "Fascinator",
            "Music",
            "Dance",
            "Maintenance and Repair",
            "Chess"
        ]

        for skill in skills:
            cursor.execute(
                "INSERT OR IGNORE INTO skills (name) VALUES (?)",
                (skill,)
            )

        comment_rows = [
            # 30–39%
            (30, 39, "This performance is well below expectation. You must take your studies more seriously."),
            (30, 39, "There is potential, but much more effort is needed. Commit to improving next term."),
            (30, 39, "You can do better. Focus, discipline, and hard work are urgently required."),
            (30, 39, "Do not give up – use this as a wake-up call to work harder and aim higher."),
            (30, 39, "Significant improvement is required. Stay determined and put in the hard work."),

            # 40–50%
            (40, 50, "This is not a true reflection of your ability. Please sit up and stay focused next term."),
            (40, 50, "You have the potential to do much better—believe in yourself and work harder next term."),
            (40, 50, "A fair attempt, but more consistency and dedication are needed. I know you can improve."),
            (40, 50, "Do not be discouraged—use this as motivation to aim higher next term. You can do it."),
            (40, 50, "Improvement is possible with better effort and time management. Stay committed and push forward."),

            # 50–60%
            (50, 60, "A satisfactory performance, but you are capable of achieving more with greater focus."),
            (50, 60, "This result shows promise—now push yourself to reach higher standards."),
            (50, 60, "You are on the right path, but stronger effort and consistency are needed."),
            (50, 60, "Average performance. Challenge yourself to improve next term."),
            (50, 60, "There is clear potential here—stay disciplined and aim for better results."),
            (50, 60, "Decent progress, but you must work harder to move beyond the basics."),

            # 60–70%
            (60, 70, "A good effort, but do not settle—keep striving for excellence."),
            (60, 70, "Solid performance, but there is room to push beyond this level."),
            (60, 70, "You are doing well—stay focused and aim even higher next term."),
            (60, 70, "A commendable result, but greater consistency will lead to stronger outcomes."),
            (60, 70, "Well done so far. Now challenge yourself to reach your full potential."),
            (60, 70, "You have built a good foundation—keep up the momentum and aim for more."),

            # 70–80%
            (70, 80, "Good work this term—now aim higher and stay consistent."),
            (70, 80, "Well done on your progress. Keep challenging yourself to go further."),
            (70, 80, "A solid result—maintain your focus and push for excellence."),
            (70, 80, "You have done well—keep up the effort and aim to improve even more."),
            (70, 80, "A commendable performance, but do not lose momentum. Stay driven."),
            (70, 80, "You are on the right path. With continued effort, even better results are within reach."),

            # 80–90%
            (80, 90, "Excellent performance—keep up the hard work and stay consistent."),
            (80, 90, "Well done! Maintain this level of focus and keep striving for excellence."),
            (80, 90, "A strong result—continue to push yourself toward the top."),
            (80, 90, "Great effort! Do not settle—aim for even greater achievement."),
            (80, 90, "You have done very well. Stay disciplined and keep progressing."),
            (80, 90, "Congratulations on your success. Now challenge yourself to reach even higher."),
        ]

        for r in comment_rows:
            cursor.execute("""
                INSERT OR IGNORE INTO principal_comments (min_average, max_average, comment)
                VALUES (?, ?, ?)
            """, r)

        # Insert sample classes
        classes = [
            ("JSS 1", "JSS"),
            ("JSS 2", "JSS"),
            ("JSS 3", "JSS"),
            ("SSS 1", "SSS"),
            ("SSS 2", "SSS"),
            ("SSS 3", "SSS")
        ]

        class_arms = [
            (1, "GOLD"), (1, "DIAMOND"),
            (2, "GOLD"), (2, "DIAMOND"),
            (3, "GOLD"), (3, "DIAMOND"),
            (4, "GOLD"), (4, "DIAMOND"),
            (5, "GOLD"), (5, "DIAMOND"),
            (6, "MASTERS")
        ]

        for name, level in classes:
            cursor.execute("INSERT OR IGNORE INTO classes (name, level) VALUES (?, ?)", (name, level))

        for class_id, arm in class_arms:
            cursor.execute("INSERT OR IGNORE INTO class_arms (class_id, arm) VALUES (?, ?)", (class_id, arm))

        # Initialize subjects and departments
        initialize_subjects_and_departments()
        
        # Initialize class subject requirements
        initialize_class_subject_requirements()

        db.commit()

        # 4. VERIFY THE DATA WAS INSERTED CORRECTLY
        print("=== Database Initialization Complete ===")

        # Check Classes
        cursor.execute("SELECT id, name, level FROM classes")
        print("Classes:")
        for row in cursor.fetchall():
            print(f"ID: {row['id']}, Name: {row['name']}, Level: {row['level']}")

        # Check Class Arms
        cursor.execute("SELECT id, class_id, arm FROM class_arms")
        print("\nClass Arms:")
        for row in cursor.fetchall():
            print(f"Arm ID: {row['id']}, Class ID: {row['class_id']}, Arm: {row['arm']}")

        # Check Departments
        cursor.execute("SELECT id, name, level FROM departments")
        print("\nDepartments:")
        for row in cursor.fetchall():
            print(f"Dept ID: {row['id']}, Name: {row['name']}, Level: {row['level']}")

        # Check Subjects
        cursor.execute("SELECT id, name, level, is_common_core FROM subjects ORDER BY level, name")
        print("\nSubjects:")
        for row in cursor.fetchall():
            print(f"Subject ID: {row['id']}, Name: {row['name']}, Level: {row['level']}, Common Core: {row['is_common_core']}")

        # Check Department-Subject relationships
        cursor.execute("""
            SELECT d.name as dept_name, s.name as subject_name, ds.is_compulsory
            FROM department_subjects ds
            JOIN departments d ON ds.department_id = d.id
            JOIN subjects s ON ds.subject_id = s.id
            ORDER BY d.name, ds.is_compulsory DESC, s.name
        """)
        print("\nDepartment-Subject Relationships:")
        for row in cursor.fetchall():
            print(f"Department: {row['dept_name']}, Subject: {row['subject_name']}, Compulsory: {row['is_compulsory']}")

        # Check Class-Subject requirements
        cursor.execute("""
            SELECT c.name as class_name, ca.arm, s.name as subject_name, csr.is_compulsory
            FROM class_subject_requirements csr
            JOIN class_arms ca ON csr.class_arm_id = ca.id
            JOIN classes c ON ca.class_id = c.id
            JOIN subjects s ON csr.subject_id = s.id
            ORDER BY c.name, ca.arm, s.name
        """)
        print("\nClass-Subject Requirements:")
        for row in cursor.fetchall():
            print(f"Class: {row['class_name']} {row['arm']}, Subject: {row['subject_name']}, Compulsory: {row['is_compulsory']}")

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get("role") not in roles:
                flash("You do not have permission to access this page", "danger")
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        return wrapper
    return decorator

def generate_reg_number(class_name, arm, session, term, index):
    """
    Generate registration number in CLASSARM-SESSION-TERM-SERIAL format.
    Example: JSS1G-2425-1-003
    """
    # Remove spaces from class name (e.g., "JSS 1" → "JSS1")
    class_abbr = class_name.replace(" ", "").upper()

    # Arm abbreviation (e.g., "Gold" → "G")
    arm_abbr = arm[0].upper() if arm else "A"

    # Convert session format "2024/2025" → "2425"
    session_parts = session.split("/")
    if len(session_parts) == 2:
        session_short = session_parts[0][-2:] + session_parts[1][:2]
    else:
        session_short = session.replace("/", "")[:4]

    return f"{class_abbr}{arm_abbr}-{session_short}-{term}-{index:03d}"


def get_subjects_by_class_arm(class_arm_id):
    """Get all subjects required for a class arm"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT s.id, s.name, s.level, s.is_common_core, csr.is_compulsory
        FROM subjects s
        JOIN class_subject_requirements csr ON s.id = csr.subject_id
        WHERE csr.class_arm_id = ?
        ORDER BY s.is_common_core DESC, s.name
    """, (class_arm_id,))
    
    return cursor.fetchall()

def get_subjects_by_department(department_id):
    """Get all subjects for a department"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT s.id, s.name, s.level, ds.is_compulsory
        FROM subjects s
        JOIN department_subjects ds ON s.id = ds.subject_id
        WHERE ds.department_id = ?
        ORDER BY ds.is_compulsory DESC, s.name
    """, (department_id,))
    
    return cursor.fetchall()

def get_class_status():
    """Get status of which classes have student data for the current year"""
    db = get_db()
    cursor = db.cursor()
    current_session = get_current_session()
    
    cursor.execute("""
        WITH ClassArmStudents AS (
            SELECT sc.class_arm_id, COUNT(DISTINCT sc.student_id) as student_count
            FROM student_classes sc
            WHERE sc.session = ?
            GROUP BY sc.class_arm_id
        )
        SELECT ca.id as arm_id, c.name as class_name, ca.arm,
               COALESCE(cas.student_count, 0) as student_count
        FROM class_arms ca
        JOIN classes c ON ca.class_id = c.id
        LEFT JOIN ClassArmStudents cas ON ca.id = cas.class_arm_id
        ORDER BY c.id, ca.arm
    """, (current_session,))
    
    classes_with_data = []
    for row in cursor.fetchall():
        classes_with_data.append({
            'arm_id': row['arm_id'],
            'class_name': row['class_name'],
            'arm': row['arm'],
            'has_data': row['student_count'] > 0,
            'student_count': row['student_count']
        })
    
    return classes_with_data

def get_current_session():
    """Get current academic session in YYYY/YYYY format"""
    current_year = datetime.now().year
    # If we're in second half of the year, current session is current_year/current_year+1
    if datetime.now().month >= 9:  # September or later
        return f"{current_year}/{current_year + 1}"
    else:
        return f"{current_year - 1}/{current_year}"

def get_current_term():
    """Get current term based on current month"""
    month = datetime.now().month
    if month in [1, 2, 3, 4]:
        return 3  # Third term
    elif month in [5, 6, 7, 8]:
        return 2  # Second term
    else:
        return 1  # First term

def initialize_subjects_and_departments():
    """Initialize subjects and departments with proper relationships"""
    db = get_db()
    cursor = db.cursor()

    # --- Define Departments ---
    departments = [
        ("Junior", "junior", "All subjects for junior classes"),
        ("Science", "senior", "Science department for senior classes"),
        ("Arts/Humanities", "senior", "Arts and Humanities department"),
        ("Commercial", "senior", "Commercial/Business department")
    ]
    
    for name, level, description in departments:
        cursor.execute("""
            INSERT OR IGNORE INTO departments (name, level, description)
            VALUES (?, ?, ?)
        """, (name, level, description))
    
    # --- Define Subjects ---
    subjects_data = [
        # Junior Subjects
        ("Mathematics", "junior", 1),
        ("English", "junior", 1),
        ("Basic Science", "junior", 0),
        ("Basic Technology", "junior", 0),
        ("Business Studies", "junior", 0),
        ("Social Studies", "junior", 0),
        ("Cultural and Creative Arts", "junior", 0),
        ("Igbo Language", "junior", 0),
        ("Yoruba Language", "junior", 0),
        ("French Language", "junior", 0),
        ("Information and Communication Technology", "junior", 0),
        ("Music", "junior", 0),
        ("History", "junior", 0),
        ("Physical and Health Education", "junior", 0),
        ("Civic Education", "junior", 1),
        ("Agricultural Science", "junior", 0),
        ("Security Education", "junior", 0),
        ("Christian Religious Studies", "junior", 0),
        ("Literature in English", "junior", 0),
        ("Diction", "junior", 0),


        # Senior Subjects (Common Core)
        ("Mathematics", "senior", 1),
        ("English", "senior", 1),
        ("Civic Education", "senior", 1),

        # Senior Science Subjects
        ("Physics", "senior", 0),
        ("Chemistry", "senior", 0),
        ("Biology", "senior", 0),
        ("Further Mathematics", "senior", 0),
        ("Agricultural Science", "senior", 0),
        ("Geography", "senior", 0),
        ("Technical Drawing", "senior", 0),
        ("Information and Communication Technology", "senior", 0),
        ("Data Processing", "senior", 0),


        # Senior Arts Subjects
        ("Literature in English", "senior", 0),
        ("Government", "senior", 0),
        ("History", "senior", 0),
        ("Christian Religious Studies", "senior", 0),
        ("Visual Arts", "senior", 0),
        ("Yoruba Language", "senior", 0),
        ("Igbo Language", "senior", 0),
        ("French Language", "senior", 0),

        # Senior Commercial Subjects
        ("Economics", "senior", 0),
        ("Commerce", "senior", 0),
        ("Financial Accounting", "senior", 0)
    ]
    
    for name, level, is_common_core in subjects_data:
        cursor.execute("""
            INSERT OR IGNORE INTO subjects (name, level, is_common_core)
            VALUES (?, ?, ?)
        """, (name, level, is_common_core))
    
    # --- Fetch IDs ---
    cursor.execute("SELECT id, name, level FROM departments")
    dept_rows = cursor.fetchall()
    departments_dict = {f"{name}_{level}": id for id, name, level in dept_rows}
    
    cursor.execute("SELECT id, name, level FROM subjects")
    subject_rows = cursor.fetchall()
    subjects_dict = {f"{name}_{level}": id for id, name, level in subject_rows}
    
    # --- Define Department-Subject Relationships ---
    # Junior General Department
    junior_dept_id = departments_dict.get("Junior_junior")
    for key, subject_id in subjects_dict.items():
        if "_junior" in key:
            cursor.execute("""
                INSERT OR IGNORE INTO department_subjects (department_id, subject_id, is_compulsory)
                VALUES (?, ?, ?)
            """, (junior_dept_id, subject_id, 1))

    # Science Department
    science_dept_id = departments_dict.get("Science_senior")
    science_subjects = [
        "Mathematics_senior", "English_senior", "Civic Education_senior",
        "Physics_senior", "Chemistry_senior", "Biology_senior",
        "Further Mathematics_senior", "Agricultural Science_senior",
        "Geography_senior", "Technical Drawing_senior"
    ]
    for key in science_subjects:
        if key in subjects_dict:
            is_compulsory = 1 if key in ["Mathematics_senior", "English_senior", "Civic Education_senior",
                                         "Physics_senior", "Chemistry_senior"] else 0
            cursor.execute("""
                INSERT OR IGNORE INTO department_subjects (department_id, subject_id, is_compulsory)
                VALUES (?, ?, ?)
            """, (science_dept_id, subjects_dict[key], is_compulsory))

    # Arts Department
    arts_dept_id = departments_dict.get("Arts/Humanities_senior")
    arts_subjects = [
        "Mathematics_senior", "English_senior", "Civic Education_senior",
        "Literature in English_senior", "Government_senior", "History_senior",
        "Christian Religious Studies_senior", "Islamic Religious Studies_senior",
        "Visual Arts_senior", "Music_senior", "Geography_senior"
    ]
    for key in arts_subjects:
        if key in subjects_dict:
            is_compulsory = 1 if key in ["Mathematics_senior", "English_senior", "Civic Education_senior",
                                         "Literature in English_senior"] else 0
            cursor.execute("""
                INSERT OR IGNORE INTO department_subjects (department_id, subject_id, is_compulsory)
                VALUES (?, ?, ?)
            """, (arts_dept_id, subjects_dict[key], is_compulsory))

    # Commercial Department
    commercial_dept_id = departments_dict.get("Commercial_senior")
    commercial_subjects = [
        "Mathematics_senior", "English_senior", "Civic Education_senior",
        "Economics_senior", "Commerce_senior", "Accounting_senior",
        "Business Studies_senior", "Financial Accounting_senior",
        "Geography_senior"
    ]
    for key in commercial_subjects:
        if key in subjects_dict:
            is_compulsory = 1 if key in ["Mathematics_senior", "English_senior",
                                         "Civic Education_senior", "Economics_senior"] else 0
            cursor.execute("""
                INSERT OR IGNORE INTO department_subjects (department_id, subject_id, is_compulsory)
                VALUES (?, ?, ?)
            """, (commercial_dept_id, subjects_dict[key], is_compulsory))

    # --- Assign Subjects to Each Class Arm ---
    cursor.execute("""
        SELECT ca.id AS id, c.name AS class_name, ca.arm AS arm, c.level AS level
        FROM class_arms ca
        JOIN classes c ON ca.class_id = c.id
    """)
    class_arms = cursor.fetchall()

    for class_arm in class_arms:
        class_arm_id = class_arm["id"]
        class_name = class_arm["class_name"].lower()

        if "jss" in class_name:
            dept_key = "Junior_junior"
        elif "sss" in class_name:
            # detect type of senior class if possible
            if "science" in class_name:
                dept_key = "Science_senior"
            elif "arts" in class_name:
                dept_key = "Arts/Humanities_senior"
            elif "commercial" in class_name:
                dept_key = "Commercial_senior"
            else:
                dept_key = "Science_senior"  # default
        else:
            continue

        dept_id = departments_dict.get(dept_key)
        if not dept_id:
            continue

        cursor.execute("""
            SELECT subject_id, is_compulsory
            FROM department_subjects
            WHERE department_id = ?
        """, (dept_id,))
        dept_subjects = cursor.fetchall()

        for subj in dept_subjects:
            cursor.execute("""
                INSERT OR IGNORE INTO class_subject_requirements (class_arm_id, subject_id, is_compulsory)
                VALUES (?, ?, ?)
            """, (class_arm_id, subj["subject_id"], subj["is_compulsory"]))

    db.commit()
    print("Subjects, departments, and class-subject links initialized successfully.")

def initialize_class_subject_requirements():
    """Set up subject requirements for each class arm"""
    db = get_db()
    cursor = db.cursor()
    
    # Get all class arms and their levels
    cursor.execute("""
        SELECT ca.id as arm_id, c.name as class_name, c.level, ca.arm
        FROM class_arms ca
        JOIN classes c ON ca.class_id = c.id
    """)
    class_arms = cursor.fetchall()
    
    for class_arm in class_arms:
        class_level = class_arm['level']  # 'JSS' or 'SSS'
        arm_id = class_arm['arm_id']
        
        if class_level == "JSS":
            # Junior classes get all junior subjects
            cursor.execute("""
                INSERT OR IGNORE INTO class_subject_requirements (class_arm_id, subject_id, is_compulsory)
                SELECT ?, s.id, 1
                FROM subjects s
                WHERE s.level = 'junior'
            """, (arm_id,))
        else:
            # Senior classes get common core subjects + department subjects
            # This will be handled when students are assigned to departments
            
            # For now, add all subjects to all senior classes
            #CHANGE LATER
            cursor.execute("""
                INSERT OR IGNORE INTO class_subject_requirements (class_arm_id, subject_id, is_compulsory)
                SELECT ?, s.id, s.is_common_core
                FROM subjects s
                WHERE s.level = 'senior'
            """, (arm_id,))
    
    db.commit()
    print("Class subject requirements initialized")


# Routes

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        db = get_db()
        cursor = db.cursor()

        cursor.execute("""
            SELECT id, username, password_hash, role
            FROM users
            WHERE username = ? AND is_active = 1
        """, (username,))
        user = cursor.fetchone()

        if not user or not check_password_hash(user["password_hash"], password):
            flash("Invalid username or password", "danger")
            return render_template("login.html")

        # ---- store session data ----
        session.clear()
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"]

        flash(f"Welcome back, {user['username']}!", "success")

        # ---- redirect based on role ----
        if user["role"] == "admin":
            return redirect(url_for("admin_dashboard"))
        elif user["role"] == "class_teacher":
            return redirect(url_for("class_teacher_home"))
        else:
            return redirect(url_for("subject_teacher_home"))

    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("You have been logged out", "info")
    return redirect(url_for("login"))

@app.route('/')
def home():
    current_session = get_current_session()
    current_term = get_current_term()
    return render_template('home.html', 
                         current_year=datetime.now().year,
                         session=current_session,
                         term=current_term)

@app.route('/uploads/photos/<filename>')
def serve_uploaded_photo(filename):
    """Serve uploaded photos from the uploads directory"""
    try:
        return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'photos'), filename)
    except FileNotFoundError:
        # Return a default image or 404 if the file doesn't exist
        return "Image not found", 404
    
# Student Biodata Management
@app.route('/manage-students')
def manage_students():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
        """)
    classes = cursor.fetchall()
    
    current_session = get_current_session()
    current_term = get_current_term()

    cursor.execute("SELECT id, name FROM departments WHERE name != 'General' ORDER BY name")
    departments = cursor.fetchall()
    current_year = datetime.now().year
    return render_template('manage_students.html', 
                           classes=classes, 
                           current_year=current_year, 
                           current_term=current_term,
                           current_session=current_session,
                           departments=departments)

@app.route('/upload-students', methods=['POST'])
def upload_students():
    class_arm_id = request.form['class_arm_id']
    session = request.form['session']  # Changed from year to session
    term = request.form['term']  # Added term

    if 'file' not in request.files:
        return redirect(url_for('manage_students'))

    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('manage_students'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        errors, success_count = process_student_upload(filepath, class_arm_id, session, term)

        if errors:
            return render_template('upload_error.html', errors=errors, success_count=success_count)

        return render_template('upload_success.html',
                              success_count=success_count,
                              message=f"{success_count} student profiles created")

    return redirect(url_for('manage_students'))

def process_student_upload(filepath, class_arm_id, session, term):
    errors = []
    success_count = 0
    db = get_db()

    try:
        # Load file
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        # Required columns
        if 'full_name' not in df.columns:
            return ["Missing required column: full_name"], 0

        cursor = db.cursor()

        # Fetch class + arm info
        cursor.execute("""
            SELECT c.id AS class_id, c.name AS class_name, c.level, a.arm
            FROM classes c
            JOIN class_arms a ON c.id = a.class_id
            WHERE a.id = ?
        """, (class_arm_id,))
        class_info = cursor.fetchone()

        if not class_info:
            return ["Invalid class arm ID"], 0

        class_name = class_info['class_name']
        class_level = class_info['level']   # JSS or SSS
        arm = class_info['arm']

        # Get existing students in this class/session/term
        cursor.execute("""
            SELECT LOWER(TRIM(s.full_name)) AS lower_name
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            WHERE sc.class_arm_id = ? AND sc.session = ? AND sc.term = ?
        """, (class_arm_id, session, term))

        existing_students = {row['lower_name'] for row in cursor.fetchall()}

        processed_names = set()
        new_students = []

        # Process rows
        for _, row in df.iterrows():
            full_name = str(row['full_name']).strip()
            if not full_name:
                continue

            lname = full_name.lower()

            if lname in processed_names:
                errors.append(f"Duplicate in file skipped: {full_name}")
                continue
            processed_names.add(lname)

            if lname in existing_students:
                errors.append(f"Already exists skipped: {full_name}")
                continue

            # Extract optional fields
            gender = None
            if 'gender' in df.columns and not pd.isna(row.get('gender', None)):
                val = str(row['gender']).strip().lower()
                if val in ['m', 'male', 'boy']: gender = "Male"
                if val in ['f', 'female', 'girl']: gender = "Female"

            dept_name = None
            if 'department' in df.columns and not pd.isna(row.get('department', None)):
                dept_name = str(row['department']).strip()

            new_students.append({
                'full_name': full_name,
                'age': int(row['age']) if 'age' in df.columns and not pd.isna(row.get('age', None)) else None,
                'gender': gender,
                'department': dept_name
            })

        if not new_students:
            return ["No new students to add."], 0

        # ===== Generate reg-number prefix =====
        # Use the same short-session rule as reg-number function
        session_parts = session.split("/")
        session_short = session_parts[0][-2:] + session_parts[1][:2]

        class_abbr = class_name.replace(" ", "").upper()
        arm_abbr = arm[0].upper()
        prefix = f"{class_abbr}{arm_abbr}-{session_short}-{term}-"

        # ===== Get highest existing index for this class-arm-session-term =====
        cursor.execute("""
            SELECT MAX(CAST(SUBSTR(reg_number, -3) AS INTEGER)) AS max_index
            FROM students
            WHERE reg_number LIKE ?
        """, (prefix + "%",))

        result = cursor.fetchone()
        start_index = (result['max_index'] + 1) if (result and result['max_index']) else 1

        # ===== Insert new students =====
        for i, student in enumerate(new_students, start=start_index):
            full_name = student['full_name']
            age = student['age']
            gender = student['gender']
            dept_name = student['department']

            # Department assignment
            department_id = None

            if class_level == "JSS":
                cursor.execute("SELECT id FROM departments WHERE name='Junior'")
                row = cursor.fetchone()
                department_id = row['id'] if row else None

            else:  # SSS student
                if dept_name:
                    dept_lower = dept_name.lower()

                    if any(k in dept_lower for k in ['sci', 'bio', 'chem', 'phy']):
                        cursor.execute("SELECT id FROM departments WHERE name='Science'")
                    elif any(k in dept_lower for k in ['art', 'human', 'lit', 'gov', 'history']):
                        cursor.execute("SELECT id FROM departments WHERE name='Arts/Humanities'")
                    elif any(k in dept_lower for k in ['comm', 'bus', 'acct', 'acc', 'eco']):
                        cursor.execute("SELECT id FROM departments WHERE name='Commercial'")
                    row = cursor.fetchone()
                    if row: department_id = row['id']

                # Default senior dept
                if not department_id:
                    cursor.execute("SELECT id FROM departments WHERE name='Science'")
                    row = cursor.fetchone()
                    department_id = row['id'] if row else None

            # ===== Generate reg number =====
            reg_number = generate_reg_number(class_name, arm, session, term, i)

            # Insert student
            cursor.execute("""
                INSERT INTO students (reg_number, full_name, age, gender, photo, department_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (reg_number, full_name, age, gender, None, department_id))

            student_id = cursor.lastrowid

            # Insert class membership
            cursor.execute("""
                INSERT OR IGNORE INTO student_classes (student_id, class_arm_id, session, term)
                VALUES (?, ?, ?, ?)
            """, (student_id, class_arm_id, session, term))

            success_count += 1

        db.commit()

    except Exception as e:
        errors.append(f"File processing error: {str(e)}")

    return errors, success_count


@app.route('/upload-student-photo/<reg_number>', methods=['POST'])
def upload_student_photo(reg_number):
    db = get_db()
    cursor = db.cursor()
    
    # Get student name for success message
    cursor.execute("SELECT full_name FROM students WHERE reg_number = ?", (reg_number,))
    student = cursor.fetchone()
    student_name = student['full_name'] if student else "Student"
    
    if 'photo' not in request.files:
        return render_template('error.html', message="No file selected")
    
    file = request.files['photo']
    if file.filename == '':
        return render_template('error.html', message="No file selected")
    
    if file and allowed_photo_file(file.filename):
        try:
            # Ensure upload directory exists
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'photos')
            os.makedirs(upload_dir, exist_ok=True)

            # Extract file extension
            file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'

            # Create filename from reg number
            filename = secure_filename(f"{reg_number}.{file_ext}")
            photo_path = os.path.join(upload_dir, filename)

            # Save original uploaded file
            file.save(photo_path)

            # Create compressed filename
            compressed_filename = f"compressed_{reg_number}.jpg"
            compressed_path = os.path.join(upload_dir, compressed_filename)

            # Compress (IMPORTANT: use the actual saved photo_path)
            compress_image(
                img_path=photo_path,
                output_path=compressed_path
            )

            # Store ONLY the compressed filename in DB
            cursor.execute(
                "UPDATE students SET photo=? WHERE reg_number=?",
                (compressed_filename, reg_number)
            )

            db.commit()
            
            # Redirect based on where the upload came from
            redirect_to = request.form.get('redirect_to', 'class_teacher')
            
            if redirect_to == 'photo_management':
                class_arm_id = request.form.get('class_arm_id')
                return redirect(url_for('student_photos') + f'?class_arm_id={class_arm_id}')
            elif redirect_to == 'class_view':
                class_arm_id = request.form.get('class_arm_id')
                session = get_current_session()
                term = get_current_term()
                return redirect(url_for('class_teacher_class_view', 
                                      class_arm_id=class_arm_id, 
                                      session=session, 
                                      term=term))
            else:
                return redirect(url_for('class_teacher_portal'))
                
        except Exception as e:
            return render_template('error.html', message=f"Error uploading photo: {str(e)}")
    
    return render_template('error.html', message="Invalid file type. Please upload JPG, JPEG, or PNG files.")

@app.route('/students')
def view_students():
    class_filter = request.args.get('class_filter', 'all')
    session_filter = request.args.get('session_filter', get_current_session())
    term_filter = request.args.get('term_filter', get_current_term())

    # Validate parameters
    if class_filter != "all":
        try:
            class_filter = int(class_filter)
        except ValueError:
            class_filter = "all"

    db = get_db()
    cursor = db.cursor()

    # Fetch all class arms for dropdown
    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """)
    classes = cursor.fetchall()

    cursor.execute("SELECT DISTINCT session FROM student_classes ORDER BY session DESC")
    sessions = [row['session'] for row in cursor.fetchall()]
    if session_filter not in sessions:
        sessions.insert(0, session_filter)

    # Build query based on filters
    if class_filter == "all":
        cursor.execute("""
            SELECT s.reg_number, s.full_name, s.age, s.photo, 
                   c.name || ' ' || a.arm AS class_name, sc.term, sc.session
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            JOIN class_arms a ON sc.class_arm_id = a.id
            JOIN classes c ON a.class_id = c.id
            WHERE sc.session = ?
            ORDER BY s.full_name
        """, (session_filter,))
    else:
       cursor.execute("""
            SELECT s.reg_number, s.full_name, s.age, s.photo, 
                   c.name || ' ' || a.arm AS class_name, sc.term, sc.session
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            JOIN class_arms a ON sc.class_arm_id = a.id
            JOIN classes c ON a.class_id = c.id
            WHERE sc.session = ? AND sc.term = ? AND sc.class_arm_id = ?
            ORDER BY s.full_name
        """, (session_filter, term_filter, class_filter))

    students = cursor.fetchall()

    return render_template('students.html',
                          students=students,
                          classes=classes,
                          sessions=sessions,
                          selected_class=class_filter,
                          selected_session=session_filter,
                          selected_term=term_filter
                          )

def process_half_term_upload(filepath, subject_id, class_arm_id, term, session):
    errors = []
    success_count = 0
    db = get_db()

    try:
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

        # df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]
        df.columns = df.columns.str.strip().str.lower()

        column_mapping = {
            'ca1': 'ca1_score',
            'ca2': 'ca2_score',
            'ca3': 'ca3_score',
            'ca4': 'ca4_score',
            'exam': 'exam_score',
            'total': 'total_score'
        }
        df.rename(columns=column_mapping, inplace=True)

        required_cols = {'full_name', 'ca1_score', 'ca2_score'}
        missing_cols = required_cols - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors, success_count

        cursor = db.cursor()
        subject_name = cursor.execute("SELECT name FROM subjects WHERE id = ?", (subject_id,)).fetchone()['name']

        for _, row in df.iterrows():
            try:
                full_name = row['full_name'].strip()
                
                # Get CA scores
                ca1_score = float(row['ca1_score']) if not pd.isna(row['ca1_score']) else 0
                ca2_score = float(row['ca2_score']) if not pd.isna(row['ca2_score']) else 0
                
                # Validate scores
                if not (0 <= ca1_score <= 5):
                    errors.append(f"Invalid CA1 score for {full_name}: {ca1_score}. Must be between 0-5")
                    continue
                if not (0 <= ca2_score <= 5):
                    errors.append(f"Invalid CA2 score for {full_name}: {ca2_score}. Must be between 0-5")
                    continue

                # Calculate total score for half-term
                total_score = ca1_score + ca2_score

                # Find student
                cursor.execute("""
                    SELECT s.id, s.department_id
                    FROM students s
                    JOIN student_classes sc ON s.id = sc.student_id
                    WHERE sc.class_arm_id = ? AND sc.session = ? AND sc.term = ? AND s.full_name = ?
                """, (class_arm_id, session, term, full_name))

                student_row = cursor.fetchone()
                if not student_row:
                    errors.append(f"Student not found in this class: {full_name}")
                    continue

                student_id = student_row['id']
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Insert or update half-term scores
                cursor.execute('''INSERT OR REPLACE INTO scores 
                                  (student_id, subject_id, class_arm_id, term, session,
                                   ca1_score, ca2_score, total_score, report_type, created_at) 
                                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                               (student_id, subject_id, class_arm_id, term, session,
                                ca1_score, ca2_score, total_score, 'half_term', now))

                success_count += 1
                
            except Exception as e:
                errors.append(f"Error processing {full_name}: {str(e)}")

        db.commit()

    except Exception as e:
        errors.append(f"File processing error: {str(e)}")

    return errors, success_count

def process_full_term_upload(filepath, subject_id, class_arm_id, term, session):
    errors = []
    success_count = 0
    db = get_db()

    try:
        # ------------------------------------------------------------------
        # 1. Read the file
        # ------------------------------------------------------------------
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath, engine='openpyxl')   # explicit engine helps

        # Normalise column names
        df.columns = df.columns.str.strip().str.lower()

        # ------------------------------------------------------------------
        # 2. Rename columns to what the code expects
        # ------------------------------------------------------------------
        column_mapping = {
            'ca1': 'ca1_score', 'ca2': 'ca2_score',
            'ca3': 'ca3_score', 'ca4': 'ca4_score',
            'exam': 'exam_score', 'total': 'total_score'
        }
        df.rename(columns=column_mapping, inplace=True)

        required_cols = {'full_name', 'ca1_score', 'ca2_score',
                         'ca3_score', 'ca4_score', 'exam_score'}
        missing_cols = required_cols - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors, success_count

        cursor = db.cursor()

        # ------------------------------------------------------------------
        # 3. Verify that the subject actually exists
        # ------------------------------------------------------------------
        cursor.execute("SELECT name FROM subjects WHERE id = ?", (subject_id,))
        subject_row = cursor.fetchone()
        if not subject_row:
            errors.append(f"Subject with id={subject_id} not found in the database.")
            return errors, success_count
        subject_name = subject_row['name']       # now safe

        
        # ------------------------------------------------------------------
        # 4. Process each row
        # ------------------------------------------------------------------
        for idx, row in df.iterrows():
            try:
                full_name = str(row['full_name']).strip()
                if not full_name or full_name.lower() == 'nan':
                    errors.append(f"Row {idx+2}: Empty or invalid full_name")
                    continue

                def read_score(val):
                    """
                    Returns:
                    - float score if valid
                    - None if blank / NaN / invalid
                    """
                    try:
                        if pd.isna(val):
                            return None
                        return float(val)
                    except:
                        return None


                ca1_score = read_score(row['ca1_score'])
                ca2_score = read_score(row['ca2_score'])
                ca3_score = read_score(row['ca3_score'])
                ca4_score = read_score(row['ca4_score'])
                exam_score = read_score(row['exam_score'])

                # ---- find the student in the correct class/term/session ----------
                cursor.execute("""
                    SELECT s.id, s.department_id
                    FROM students s
                    JOIN student_classes sc ON s.id = sc.student_id
                    WHERE sc.class_arm_id = ? 
                      AND sc.session = ? 
                      AND sc.term = ? 
                      AND LOWER(TRIM(s.full_name)) = LOWER(TRIM(?))
                """, (class_arm_id, session, term, full_name))

                student_row = cursor.fetchone()
                if not student_row:
                    errors.append(f"Student not enrolled in this class/term/session: {full_name}")
                    continue

                student_id = student_row['id']

                # ---- insert / replace the score ----------------------------------

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if ca1_score and ca2_score and ca3_score and ca4_score and exam_score:

                    # ---- validation --------------------------------------------------
                    if not (0 <= ca1_score <= 5):
                        errors.append(f"{full_name}: CA1 {ca1_score} (must be 0-5)")
                        continue
                    if not (0 <= ca2_score <= 5):
                        errors.append(f"{full_name}: CA2 {ca2_score} (must be 0-5)")
                        continue
                    if not (0 <= ca3_score <= 5):
                        errors.append(f"{full_name}: CA3 {ca3_score} (must be 0-5)")
                        continue
                    if not (0 <= ca4_score <= 5):
                        errors.append(f"{full_name}: CA4 {ca4_score} (must be 0-5)")
                        continue
                    if not (0 <= exam_score <= 80):
                        errors.append(f"{full_name}: Exam {exam_score} (must be 0-80)")
                        continue
                    
                    total_score = ca1_score + ca2_score + ca3_score + ca4_score + exam_score

                    cursor.execute("""
                        INSERT OR REPLACE INTO scores 
                        (student_id, subject_id, class_arm_id, term, session,
                        ca1_score, ca2_score, ca3_score, ca4_score,
                        exam_score, total_score, report_type, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (student_id, subject_id, class_arm_id, term, session,
                        ca1_score, ca2_score, ca3_score, ca4_score,
                        exam_score, total_score, 'full_term', now))

                    success_count += 1

            except Exception as row_err:
                errors.append(f"Row {idx+2} ({full_name}): {str(row_err)}")

        db.commit()

    except Exception as e:
        errors.append(f"File processing error: {str(e)}")

    return errors, success_count

@app.route('/upload-half-term-results')
def upload_half_term_results():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM subjects ORDER BY name")
    subjects = cursor.fetchall()

    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """)
    classes = cursor.fetchall()

    current_session = get_current_session()

    return render_template('upload_half_term_results.html',
                          subjects=subjects,
                          classes=classes,
                          current_session=current_session)

@app.route('/upload-full-term-results')
def upload_full_term_results():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM subjects ORDER BY name")
    subjects = cursor.fetchall()

    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """)
    classes = cursor.fetchall()

    current_session = get_current_session()

    return render_template('upload_full_term_results.html',
                          subjects=subjects,
                          classes=classes,
                          current_session=current_session)


    subject_id = request.form['subject_id']
    class_arm_id = request.form['class_arm_id']
    term = request.form['term']
    session = request.form['session']

    if 'file' not in request.files:
        return redirect(url_for('upload_full_term_results'))

    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('upload_full_term_results'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        errors, success_count = process_full_term_upload(filepath, subject_id, class_arm_id, term, session)

        if errors:
            return render_template('upload_error.html', errors=errors, success_count=success_count)

        return render_template('upload_success.html',
                              success_count=success_count,
                              message=f"{success_count} full-term scores uploaded")

    return redirect(url_for('upload_full_term_results'))

@app.route('/preview-results', methods=['POST'])
def preview_results():
    """Show preview of uploaded result file before confirming."""
    try:
        report_type = request.form.get("report_type")  # 'half_term' or 'full_term'
        subject_id = request.form.get("subject_id")
        class_arm_id = request.form.get("class_arm_id")
        term = int(request.form.get("term"))
        session = request.form.get("session")

        # Validate form inputs
        if not all([report_type, subject_id, class_arm_id, term, session]):
            return render_template("error.html", message="Missing required fields.")

        # Validate file
        if 'file' not in request.files:
            return render_template("error.html", message="No file uploaded.")

        file = request.files['file']
        if file.filename == '':
            return render_template("error.html", message="Empty file name.")

        # Save file temporarily
        temp_filename = f"temp_{datetime.now().timestamp()}_{secure_filename(file.filename)}"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        file.save(temp_path)

        # Load DataFrame
        df = pd.read_excel(temp_path) if temp_path.endswith(".xlsx") else pd.read_csv(temp_path)
        df.columns = df.columns.str.strip().str.lower()

        # Column requirements
        required_cols = {'full_name', 'ca1', 'ca2'} if report_type == "half_term" else {
            'full_name', 'ca1', 'ca2', 'ca3', 'ca4', 'exam'
        }

        # Detect missing columns
        missing = required_cols - set(df.columns)
        if missing:
            return render_template(
                "upload_error.html",
                errors=[f"Missing required columns: {', '.join(missing)}"],
                success_count=0
            )

        # Check existing scores that might be overwritten
        db = get_db()
        cursor = db.cursor()
        overwrite_warnings = []

        cursor.execute("""
            SELECT sub.name
            FROM subjects sub
            WHERE sub.id=?
        """, (subject_id,)
        )
        subject = cursor.fetchone()

        cursor.execute("""
            SELECT a.id AS arm_id, c.name AS class_name, a.arm
            FROM class_arms a
            JOIN classes c ON a.class_id = c.id
            WHERE a.id = ?
        """, (class_arm_id,))
        class_data = cursor.fetchone()

        for name in df['full_name']:

            # 1. Lookup student ID from name
            cursor.execute("""
                SELECT s.id
                FROM students s
                JOIN student_classes sc ON s.id = sc.student_id
                WHERE sc.class_arm_id=? AND sc.session=? AND sc.term=? AND s.full_name=?
            """, (class_arm_id, session, term, name.strip()))

            student = cursor.fetchone()
            if not student:
                continue  # or collect as an error
            
            student_id = student['id']

            # 2. Check if scores already exist for this student
            cursor.execute("""
                SELECT 1 FROM scores
                WHERE student_id=? AND subject_id=? AND class_arm_id=? AND term=? AND session=?
            """, (student_id, subject_id, class_arm_id, term, session))

            if cursor.fetchone():
                overwrite_warnings.append(name.strip())

        return render_template(
            "results_preview.html",
            df=df.to_dict(orient='records'),
            report_type=report_type,
            subject=subject,
            class_data=class_data,
            subject_id=subject_id,    
            class_arm_id=class_arm_id,
            term=term,
            session=session,
            temp_path=temp_path,
            overwrite_warnings=overwrite_warnings
        )

    except Exception as e:
        return render_template("error.html", message=str(e))

@app.route('/confirm-results-upload', methods=['POST'])
def confirm_results_upload():
    """Final commit of results (after preview)."""
    report_type = request.form.get("report_type")
    subject_id = request.form.get("subject_id")
    class_arm_id = request.form.get("class_arm_id")
    term = request.form.get("term")
    session = request.form.get("session")
    temp_path = request.form.get("temp_path")

    if not os.path.exists(temp_path):
        return render_template("error.html", message="Temporary file missing. Please re-upload.")

    if report_type not in ("half_term", "full_term"):
        return render_template("error.html", message="Invalid report type.")

    # Call correct processing function
    if report_type == "half_term":
        errors, success_count = process_half_term_upload(temp_path, subject_id, class_arm_id, term, session)
    else:
        errors, success_count = process_full_term_upload(temp_path, subject_id, class_arm_id, term, session)

    # Cleanup
    os.remove(temp_path)

    if errors:
        return render_template("upload_error.html", errors=errors, success_count=success_count)

    return render_template("upload_success.html",
                           message=f"{success_count} records uploaded!",
                           success_count=success_count)

@app.route('/results')
def view_results():
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('''
        SELECT s.reg_number, s.full_name, 
               c.name || ' - ' || ca.arm AS class_name,  -- Combine class name and arm
               sub.name AS subject, sc.total_score, sc.term, sc.session
        FROM scores sc
        JOIN students s ON sc.student_id = s.id
        JOIN subjects sub ON sc.subject_id = sub.id
        JOIN class_arms ca ON sc.class_arm_id = ca.id
        JOIN classes c ON ca.class_id = c.id
        ORDER BY sc.session DESC, sc.term DESC, c.name, ca.arm, s.full_name
    ''')
    results = cursor.fetchall()
    return render_template('results.html', results=results)

def get_best_student_match(class_arm_id, session, full_name):
    """Simple approach that always returns one student or None"""
    search_name = full_name.strip()
    
    db = get_db()
    cursor = db.cursor()

    # Get all students in the class
    cursor.execute("""
        SELECT s.*, c.name || ' ' || a.arm AS class_name
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        JOIN class_arms a ON sc.class_arm_id = a.id
        JOIN classes c ON a.class_id = c.id
        WHERE sc.class_arm_id = ? AND sc.session = ?
    """, (class_arm_id, session))
    
    all_students = cursor.fetchall()
    
    if not all_students:
        return None
    
    # Convert to list of dictionaries
    students = [dict(student) for student in all_students]
    
    search_lower = search_name.lower()
    
    # Try exact match first
    for student in students:
        if student['full_name'].lower() == search_lower:
            return student
    
    # Try contains match
    contains_matches = [s for s in students if search_lower in s['full_name'].lower()]
    if len(contains_matches) == 1:
        return contains_matches[0]
    
    # Try reverse contains (search term contains student name)
    reverse_matches = [s for s in students if s['full_name'].lower() in search_lower]
    if len(reverse_matches) == 1:
        return reverse_matches[0]
    
    # Try word-based matching
    search_words = set(search_lower.split())
    word_matches = []
    for student in students:
        student_words = set(student['full_name'].lower().split())
        if search_words.intersection(student_words):
            word_matches.append(student)
    
    if len(word_matches) == 1:
        return word_matches[0]
    
    return None

@app.route('/generate-reports', methods=['GET', 'POST'])
def generate_reports():
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        class_arm_id = request.form.get('class_arm_id')
        term = int(request.form['term'])
        session = request.form['session']
        report_type = request.form.get('report_type', 'full_term')
        
        cursor.execute("SELECT class_id FROM class_arms WHERE id = ?", (class_arm_id,))
        class_result = cursor.fetchone()
        class_id = class_result['class_id']

        cursor.execute("SELECT level FROM classes WHERE id = ?", (class_id,))
        level = cursor.fetchone()
        class_level = level['level']

        results = []

        results = cursor.execute("""
                SELECT s.id, s.full_name, s.reg_number, c.name || ' ' || a.arm AS class_name,
                    AVG(sc.total_score) AS average
                FROM students s
                JOIN scores sc ON s.id = sc.student_id
                JOIN student_classes x ON s.id = x.student_id
                JOIN class_arms a ON x.class_arm_id = a.id
                JOIN classes c ON a.class_id = c.id
                WHERE a.class_id = ? AND x.term = ? AND x.session = ? AND sc.report_type = ?
                GROUP BY s.id
                ORDER BY average DESC
            """, (class_id, term, session, report_type)).fetchall()

        class_avg = sum(r["average"] for r in results)/len(results) if results else 0

        # Convert to list of tuples (student_id, average)
        entries = [(row['id'], row['average']) for row in results]

        # Sort by highest average
        entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)

        # Assign positions (1st, 2nd, 3rd…)
        rankings = {}
        position = 1

        for i, (student_id, avg) in enumerate(entries_sorted):
            # Handle ties: if same score as previous student, same rank
            if i > 0 and avg == entries_sorted[i - 1][1]:
                rankings[student_id] = rankings[entries_sorted[i - 1][0]]
            else:
                rankings[student_id] = position
            
            position += 1

        # --- Single Student Report ---
        if full_name:
            student = get_best_student_match(class_arm_id, session, full_name)
            print(student)

            if not student:
                return render_template("error.html", message="No student found")

            cursor.execute("""
                    SELECT sub.name AS subject, 
                           sc.ca1_score, sc.ca2_score, sc.ca3_score, sc.ca4_score, 
                           sc.exam_score, sc.total_score, sc.report_type
                    FROM scores sc
                    JOIN subjects sub ON sc.subject_id = sub.id
                    WHERE sc.student_id = ? AND sc.term = ? AND sc.session = ? AND sc.report_type = ?
                """, (student["id"], term, session, report_type))
            scores = cursor.fetchall()

            if not scores:
                return render_template("error.html", message="No results found")

            cursor.execute("""
                    SELECT days_present, days_absent, days_late, total_school_days
                    FROM attendance_summary 
                    WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
                """, (student['id'], class_arm_id, term, session))
            attendance_summary = cursor.fetchone()

            cursor.execute("""
                SELECT handwriting, sports_participation, practical_skills,
                    punctuality, politeness, neatness,
                    class_teacher_comment, principal_comment
                FROM student_assessments
                WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
            """, (student["id"], class_arm_id, term, session))
            assessment = cursor.fetchone()

            cursor.execute('''SELECT sk.name, ss.score
                FROM student_skills ss
                JOIN skills sk ON ss.skill_id = sk.id
                WHERE ss.student_id = ? AND ss.term = ? AND ss.session = ?''', (student["id"], term, session))
            skills = cursor.fetchone()

            total = sum(r["total_score"] for r in scores)
            average = total / len(scores)

            student_position = rankings.get(student['id'], None)

            if class_level == "JSS":
                position = student_position
                grade = None
            else:
                position = None
                grade = grade_from_average(average)


            # Choose template based on report type
            template_name = "half_term_report.html" if report_type == "half_term" else "full_term_report.html"
            logo_path = os.path.join(app.root_path, 'static', 'kembos_logo_nobg.png')


            return render_template(template_name,
                                           student=student,
                                           class_name=student["class_name"],
                                           scores=scores,
                                           term=term,
                                           logo_path=logo_path,
                                           session=session,
                                           average=average,
                                           class_average=class_avg,
                                           position=position,
                                           grade=grade,
                                           report_type=report_type,
                                           skills=skills,
                                           attendance_summary=attendance_summary,
                                           assessment=assessment,
                                           year=datetime.now().year,
                                           current_date=datetime.now().strftime("%Y-%m-%d"))

       # --- Whole Class Batch Reports ---
        cursor.execute("""
            SELECT s.id, s.reg_number, s.full_name, s.age, s.photo, 
                c.name || ' ' || a.arm AS class_name
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            JOIN class_arms a ON sc.class_arm_id = a.id
            JOIN classes c ON a.class_id = c.id
            WHERE sc.class_arm_id = ? AND sc.session = ?
            ORDER BY s.full_name
        """, (class_arm_id, session))
        students = cursor.fetchall()

        if not students:
            return render_template("error.html", message="No students found for this class/year")

        report_type = request.form.get("report_type", "full_term")

        # Sanitize ZIP filename
        safe_session = session.replace("/", "_")
        zip_filename = f"class_{class_arm_id}_term{term}_{safe_session}_reports.zip"

        # Prepare in-memory ZIP buffer
        zip_buffer = BytesIO()

        # Detect your current host (local or LAN)
        host_url = request.host_url.rstrip('/')

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for student in students:
                    reg_number = student["reg_number"]
                    
                    # Construct the exact URL for the student's report page
                    report_url = f"{host_url}/preview-report?student_id={student['id']}&term={term}&session={session}&report_type={report_type}"
                    
                    pdf_filename = f"{student['full_name'].replace(' ', '_')}_report.pdf"
                    pdf_path = os.path.join(tempfile.gettempdir(), pdf_filename)
                    
                    page = context.new_page()
                    print(f"🧾 Generating PDF for {student['full_name']} ({report_type})")
                    
                    try:
                        page.goto(report_url, wait_until="networkidle")
                        page.pdf(
                            path=pdf_path,
                            format="A4",
                            print_background=True,
                            margin={"top": "10mm", "bottom": "10mm", "left": "10mm", "right": "10mm"}
                        )
                        zf.write(pdf_path, pdf_filename)
                    except Exception as e:
                        print(f"❌ Error generating report for {student['full_name']}: {e}")
                    finally:
                        if os.path.exists(pdf_path):
                            os.remove(pdf_path)
                        page.close()
            
            browser.close()

        # Return ZIP as downloadable file
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype="application/zip"
        )


    # --- GET request → show report form ---
    cursor.execute("""
                SELECT a.id AS arm_id, c.name AS class_name, a.arm
                FROM class_arms a
                JOIN classes c ON a.class_id = c.id
                ORDER BY c.id, a.arm
            """)
    classes = cursor.fetchall()
    current_session=get_current_session()
    return render_template("report_form.html", classes=classes, current_session=current_session)

def compress_image(img_path, output_path, max_width=300, quality=70):
    img = Image.open(img_path)

    # --- FIX ORIENTATION ---
    try:
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == 'Orientation':
                break

        exif = img._getexif()
        if exif is not None:
            orientation_value = exif.get(orientation)
            if orientation_value == 3:
                img = img.rotate(180, expand=True)
            elif orientation_value == 6:
                img = img.rotate(270, expand=True)
            elif orientation_value == 8:
                img = img.rotate(90, expand=True)
    except Exception:
        # If EXIF is missing or unreadable, skip orientation correction
        pass

    # --- RESIZE AND COMPRESS ---
    img.thumbnail((max_width, max_width), Image.LANCZOS)
    rgb_img = img.convert('RGB')
    rgb_img.save(output_path, format="JPEG", optimize=True, quality=quality)

    return output_path

def grade_from_average(avg):
    if avg >= 75:
        return "A+"
    elif avg >= 65:
        return "B+"
    elif avg >= 60:
        return "C+"

@app.route("/preview-report")
def preview_report():
    student_id = request.args.get("student_id")
    term = request.args.get("term")
    session = request.args.get("session")
    report_type = request.args.get("report_type", "full_term")

    if not student_id or not term or not session:
        return "Missing parameters", 400

    db = get_db()
    cursor = db.cursor()

    # Get student info
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.age, s.photo, 
               c.name || ' ' || a.arm AS class_name, s.department_id, s.gender
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        JOIN class_arms a ON sc.class_arm_id = a.id
        JOIN classes c ON a.class_id = c.id
        WHERE s.id = ? AND sc.session = ?
    """, (student_id, session))
    student = cursor.fetchone()
    if not student:
        return "Student not found", 404

    # Get scores
    cursor.execute("""
        SELECT sub.name AS subject, 
               sc.ca1_score, sc.ca2_score, sc.ca3_score, sc.ca4_score,
               sc.exam_score, sc.total_score
        FROM scores sc
        JOIN subjects sub ON sc.subject_id = sub.id
        WHERE sc.student_id = ? AND sc.term = ? AND sc.session = ? AND sc.report_type = ?
    """, (student_id, term, session, report_type))
    scores = cursor.fetchall()

    # Get attendance
    cursor.execute("""
        SELECT days_present, days_absent, days_late, total_school_days
        FROM attendance_summary 
        WHERE student_id = ? AND class_arm_id = (
            SELECT class_arm_id FROM student_classes 
            WHERE student_id = ? AND session = ?
        )
        AND term = ? AND session = ?
    """, (student_id, student_id, session, term, session))
    attendance_summary = cursor.fetchone()

    cursor.execute("""
        SELECT handwriting, sports_participation, practical_skills,
            punctuality, politeness, neatness,
            class_teacher_comment, principal_comment
        FROM student_assessments
        WHERE student_id = ? AND term = ? AND session = ?
    """, (student["id"], term, session))
    assessment = cursor.fetchone()

    cursor.execute('''SELECT sk.name, ss.score
                FROM student_skills ss
                JOIN skills sk ON ss.skill_id = sk.id
                WHERE ss.student_id = ? AND ss.term = ? AND ss.session = ?''', (student["id"], term, session))
    skills = cursor.fetchone()
    
    cursor.execute("""
        SELECT class_arm_id
        FROM student_classes
        WHERE student_id = ?
    """, (student["id"],))
    class_arm = cursor.fetchone()

    cursor.execute("SELECT class_id FROM class_arms WHERE id = ?", (class_arm['class_arm_id'],))
    class_result = cursor.fetchone()
    class_id = class_result['class_id']

    cursor.execute("SELECT level FROM classes WHERE id = ?", (class_id,))
    level = cursor.fetchone()
    class_level = level['level']

    results = []

    results = cursor.execute("""
                SELECT s.id, s.full_name, s.reg_number, c.name || ' ' || a.arm AS class_name,
                    AVG(sc.total_score) AS average
                FROM students s
                JOIN scores sc ON s.id = sc.student_id
                JOIN student_classes x ON s.id = x.student_id
                JOIN class_arms a ON x.class_arm_id = a.id
                JOIN classes c ON a.class_id = c.id
                WHERE a.class_id = ? AND x.term = ? AND x.session = ? AND sc.report_type = ?
                GROUP BY s.id
                ORDER BY average DESC
            """, (class_id, term, session, report_type)).fetchall()

    # Convert to list of tuples (student_id, average)
    entries = [(row['id'], row['average']) for row in results]

    # Sort by highest average
    entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)

    # Assign positions (1st, 2nd, 3rd…)
    rankings = {}
    position = 1

    for i, (student_id, avg) in enumerate(entries_sorted):
        # Handle ties: if same score as previous student, same rank
        if i > 0 and avg == entries_sorted[i - 1][1]:
            rankings[student_id] = rankings[entries_sorted[i - 1][0]]
        else:
            rankings[student_id] = position
            
        position += 1

    class_avg = sum(r["average"] for r in results)/len(results) if results else 0
    total = sum(r["total_score"] or 0 for r in scores) if scores else 0
    average = total / len(scores) if scores else 0
    student_position = rankings.get(student["id"], None)

    if class_level == "JSS":
        position = student_position
        grade = None
    else:
        position = None
        grade = grade_from_average(average)

    template_name = "full_term_report.html" if report_type == "full_term" else "half_term_report.html"

    return render_template(template_name,
                           student=student,
                           class_name=student["class_name"],
                           scores=scores,
                           term=term,
                           session=session,
                           class_average=class_avg,
                           position=position,
                           grade=grade,
                           skills=skills,
                           average=average,
                           report_type=report_type,
                           attendance_summary=attendance_summary,
                           assessment=assessment,
                           year=datetime.now().year,
                           current_date=datetime.now().strftime("%Y-%m-%d"))

@app.route('/download-student-template')
def download_student_template():
    """
    Download Excel template for class teachers to upload student biodata.
    """
    sample_data = {
        'full_name': ['John Doe', 'Jane Smith', 'Mike Johnson'],
        'age': [15, 16, 14],
        'gender': ['Male', 'Female', 'Male'],
        'department': ['Science', 'Arts', 'Commercial']
    }
    df = pd.DataFrame(sample_data)

    # Write to Excel (BytesIO for in-memory download)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Student_Data', index=False)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='student_upload_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download-result-template')
def download_result_template():
    """Download Excel template for result uploads (auto formulas + locked names)."""
    class_arm_id = request.args.get('class_arm_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    term = request.args.get('term')
    session = request.args.get('session')

    if not all([class_arm_id, subject_id, term, session]):
        abort(400, description="Missing required parameters")

    db = get_db()
    cursor = db.cursor()

    # --- Subject & Class Info ---
    cursor.execute("""
        SELECT s.name AS subject_name, c.name AS class_name, a.arm
        FROM subjects s
        JOIN class_arms a ON a.id = ?
        JOIN classes c ON a.class_id = c.id
        WHERE s.id = ?
    """, (class_arm_id, subject_id))
    info = cursor.fetchone()
    if not info:
        abort(404, description="Invalid class or subject selection")

    # --- Get Students ---
    cursor.execute("""
        SELECT st.full_name
        FROM students st
        JOIN student_classes sc ON st.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ? AND sc.term = ?
        ORDER BY st.full_name
    """, (class_arm_id, session, term))
    students = cursor.fetchall()
    if not students:
        abort(404, description="No students found for this class/session")

    # --- Create DataFrame ---
    data = {
        'full_name': [s['full_name'] for s in students],
        'CA1': [''] * len(students),
        'CA2': [''] * len(students),
        'CA3': [''] * len(students),
        'CA4': [''] * len(students),
        'CA Total': [''] * len(students),
        'Exam': [''] * len(students),
        'Total': [''] * len(students)
    }
    df = pd.DataFrame(data)

    # --- Write to Excel ---
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Results', index=False)
    output.seek(0)

    # --- Load workbook for styling, formulas, and protection ---
    wb = load_workbook(output)
    ws = wb.active

    # --- Header Styling ---
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Insert Formulas ---
    ca_total_col = 6  # F
    exam_col = 7      # G
    total_col = 8     # H
    num_students = len(students)

    for i in range(2, num_students + 2):
        # CA Total = SUM(B:E)
        ws.cell(row=i, column=ca_total_col).value = f"=SUM(B{i}:E{i})"
        # Total = F + G
        ws.cell(row=i, column=total_col).value = f"=F{i}+G{i}"

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto column width ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # --- Protect sheet ---
    ws.protection.sheet = True
    ws.protection.password = "kembos"  # You can change this password if you want

    # Make only "Full Name" column locked, others editable
    for row in ws.iter_rows(min_row=2, max_row=num_students + 1, min_col=1, max_col=8):
        for cell in row:
            if cell.column == 1:  # Full Name
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # --- Save to memory ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"result_template_{info['class_name']}_{info['arm']}_{info['subject_name']}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename.replace(" ", "_"),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def get_filtered_performance_base(class_arm_id, term, session, report_type):
    # Returns (cte_sql, params)
    parts = []
    params = []

    if class_arm_id:
        parts.append("x.class_arm_id = ?")
        params.append(class_arm_id)
    if term:
        parts.append("x.term = ?")
        params.append(int(term))
    if session:
        parts.append("x.session = ?")
        params.append(session)
    if report_type:
        parts.append("sc.report_type = ?")
        params.append(report_type)

    where = " AND ".join(parts) if parts else "1 = 1"

    cte = f"""
    WITH filtered_performance AS (
        SELECT
            s.id                AS student_id,
            s.full_name,
            s.gender,
            s.reg_number,
            x.class_arm_id,
            c.name || ' ' || a.arm  AS class_name,
            sc.subject_id,
            sub.name            AS subject_name,
            sub.level,
            sc.total_score,
            sc.approved,
            AVG(sc.total_score) OVER (PARTITION BY s.id) AS student_avg,   -- optional window
            sc.total_score
        FROM students s
        JOIN student_classes     x  ON s.id = x.student_id
        JOIN class_arms          a  ON x.class_arm_id = a.id
        JOIN classes             c  ON a.class_id = c.id
        JOIN scores              sc ON s.id = sc.student_id AND x.class_arm_id = sc.class_arm_id
        JOIN subjects            sub ON sc.subject_id = sub.id
        WHERE {where}
    )
    """

    return cte, params

@app.route("/admin-dashboard", methods=["GET"])
@login_required
@roles_required("admin")
def admin_dashboard():
    db = get_db()
    cursor = db.cursor()

    # Get filter parameters (all optional)
    class_arm_id = request.args.get("class_arm_id")
    term = request.args.get("term")
    session = request.args.get("session")
    report_type = request.args.get("report_type", "full_term")

    cte, params = get_filtered_performance_base(class_arm_id, term, session, report_type)

    # Fetch all class arms for dropdown/filter
    classes = cursor.execute("""
        SELECT a.id as arm_id, c.name as class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """).fetchall()

    # Common base filters (for student_classes)
    base_filters = []
    base_params = []
    if class_arm_id:
        base_filters.append("x.class_arm_id = ?")
        base_params.append(class_arm_id)
    if term:
        base_filters.append("x.term = ?")
        base_params.append(term)
    if session:
        base_filters.append("x.session = ?")
        base_params.append(session)

    base_where = " AND ".join(base_filters) if base_filters else "1=1"
    base_params_list = base_params[:]  # copy

    # For queries that need report_type (all scores-based queries)
    report_where = f"sc.report_type = ?" if report_type else "1=1"
    report_params = [report_type] if report_type else []

    # 1. Student performance query
    student_query = f"""
        SELECT 
            s.id, s.full_name, s.reg_number, s.gender,
            c.name || ' ' || a.arm AS class_name,
            AVG(sc.total_score) AS average,
            SUM(CASE WHEN sc.approved = 1 THEN 1 ELSE 0 END) AS approved_count,
            COUNT(sc.id) AS subject_count
        FROM students s
        JOIN scores sc ON s.id = sc.student_id
        JOIN student_classes x ON s.id = x.student_id AND sc.class_arm_id = x.class_arm_id
        JOIN class_arms a ON x.class_arm_id = a.id
        JOIN classes c ON a.class_id = c.id
        WHERE {base_where} AND {report_where}
        GROUP BY s.id
        HAVING subject_count > 0
        ORDER BY average DESC
    """
    students = cursor.execute(student_query, base_params_list + report_params).fetchall()

    # 2. Class averages
    class_avg_query = f"""
        SELECT 
            c.name || ' ' || a.arm AS class_name,
            COUNT(DISTINCT x.student_id) AS student_count,
            AVG(sc.total_score) AS class_average
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        JOIN student_classes x ON a.id = x.class_arm_id
        JOIN scores sc ON x.student_id = sc.student_id AND x.class_arm_id = sc.class_arm_id
        WHERE {base_where} AND {report_where}
        GROUP BY a.id, c.name, a.arm
        ORDER BY class_average DESC
    """
    class_averages = cursor.execute(class_avg_query, base_params_list + report_params).fetchall()

    # 3. Gender performance
    gender_query = f"""
        SELECT 
            s.gender,
            AVG(sc.total_score) AS avg_score,
            COUNT(DISTINCT s.id) AS student_count
        FROM students s
        JOIN scores sc ON s.id = sc.student_id
        JOIN student_classes x ON s.id = x.student_id AND sc.class_arm_id = x.class_arm_id
        WHERE {base_where} AND {report_where}
        GROUP BY s.gender
        ORDER BY avg_score DESC
    """
    gender_performance = cursor.execute(gender_query, base_params_list + report_params).fetchall()

    
    # 4. Overall stats (using the fixed CTE version from previous response)
    # overall_query = f"""
    # WITH student_averages AS (
    #     SELECT 
    #         s.id,
    #         AVG(sc.total_score) AS avg_score
    #     FROM students s
    #     JOIN scores sc ON s.id = sc.student_id
    #     JOIN student_classes x ON s.id = x.student_id AND sc.class_arm_id = x.class_arm_id
    #     WHERE {base_where} AND {report_where}
    #     GROUP BY s.id
    #     HAVING COUNT(sc.id) > 0
    # )
    # SELECT 
    #     COUNT(*) AS total_students,
    #     AVG(avg_score) AS school_average,
    #     SUM(CASE WHEN avg_score >= 70 THEN 1 ELSE 0 END) AS above_70,
    #     SUM(CASE WHEN avg_score < 70 THEN 1 ELSE 0 END) AS below_70,
    #     (SELECT COUNT(*) FROM scores sc WHERE {base_where} AND sc.report_type = ? AND sc.approved = 1) AS approved_scores,
    #     (SELECT COUNT(*) FROM scores sc WHERE {base_where} AND sc.report_type = ?) AS total_scores
    # FROM student_averages;
    # """
    overall = cursor.execute(f"""
        {cte}
        SELECT
            COUNT(DISTINCT student_id)                  AS total_students,
            AVG(student_avg)                            AS school_average,
            SUM(CASE WHEN student_avg >= 70 THEN 1 ELSE 0 END) AS above_70,
            SUM(CASE WHEN student_avg <  70 THEN 1 ELSE 0 END) AS below_70
        FROM filtered_performance
        """, params).fetchone()

    # 5. Prepare chart data
    # Top 10 students (for bar/line chart)
    top_students_data = {
        "labels": [s["full_name"] for s in students[:10]],
        "values": [round(s["average"], 2) for s in students[:10]]
    }

    # Class averages chart data
    class_avg_data = {
        "labels": [row["class_name"] for row in class_averages],
        "values": [round(row["class_average"], 2) for row in class_averages],
        "counts": [row["student_count"] for row in class_averages]  # optional for tooltip
    }

    # Benchmark distribution (for pie/donut chart)
    benchmark_data = {
        "labels": ["Above 70%", "Below 70%"],
        "values": [overall["above_70"], overall["below_70"]]
    } if overall else {"labels": [], "values": []}

    return render_template(
        "admin_dashboard.html",
        classes=classes,                    # for dropdown
        students=students,                  # full list or paginated
        class_averages=class_averages,      # table of class averages
        overall_stats=overall,        # KPIs: total students, school avg, etc.
        gender_performance=gender_performance,
        term=term,
        session=session,
        report_type=report_type,
        # Chart data (pass to JavaScript, e.g., Chart.js)
        top_students_data=top_students_data,
        class_avg_data=class_avg_data,
        benchmark_data=benchmark_data,
        # Optional: add more chart data here
    )

@app.route("/admin/students")
def admin_students():
    db = get_db()
    cursor = db.cursor()
    
    search = request.args.get("search", "")
    
    if search:
        cursor.execute("""
            SELECT s.*, d.name AS dept_name
            FROM students s
            LEFT JOIN departments d ON d.id = s.department_id
            WHERE s.full_name LIKE ? OR s.reg_number LIKE ?
            ORDER BY s.full_name
        """, (f"%{search}%", f"%{search}%"))
    else:
        cursor.execute("""
            SELECT s.*, d.name AS dept_name
            FROM students s
            LEFT JOIN departments d ON d.id = s.department_id
            ORDER BY s.full_name
        """)
    
    students = cursor.fetchall()
    
    return render_template("admin/students_list.j2", students=students, search=search)

@app.route("/admin/student/<int:student_id>/edit")
def edit_student(student_id):
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT * FROM students WHERE id=?
    """, (student_id,))
    student = cursor.fetchone()
    
    cursor.execute("SELECT id, name FROM departments")
    departments = cursor.fetchall()
    
    return render_template("admin/edit_student.j2",
                           student=student,
                           departments=departments)

@app.route("/admin/student/<int:student_id>/update", methods=["POST"])
def update_student(student_id):
    db = get_db()
    cursor = db.cursor()
    
    full_name = request.form.get("full_name").strip()
    age = request.form.get("age")
    gender = request.form.get("gender")
    department_id = request.form.get("department_id")
    photo = request.files.get("photo")
    
    # Update photo if uploaded
    photo_filename = None
    if photo and photo.filename:
        from PIL import Image, ExifTags
        import io, base64, os
        from werkzeug.utils import secure_filename
        
        upload_dir = os.path.join(app.config["UPLOAD_FOLDER"], "photos")
        os.makedirs(upload_dir, exist_ok=True)
        
        ext = photo.filename.rsplit(".", 1)[-1].lower()
        saved_name = secure_filename(f"student_{student_id}.{ext}")
        saved_path = os.path.join(upload_dir, saved_name)
        photo.save(saved_path)
        
        # Compress
        compressed_name = f"compressed_{saved_name}"
        compressed_path = os.path.join(upload_dir, compressed_name)
        compress_image(saved_path, compressed_path)
        
        photo_filename = compressed_name
    
    # Update DB
    if photo_filename:
        cursor.execute("""
            UPDATE students
            SET full_name=?, age=?, gender=?, department_id=?, photo=?
            WHERE id=?
        """, (full_name, age, gender, department_id, photo_filename, student_id))
    else:
        cursor.execute("""
            UPDATE students
            SET full_name=?, age=?, gender=?, department_id=?
            WHERE id=?
        """, (full_name, age, gender, department_id, student_id))
    
    db.commit()
    
    return redirect(url_for("admin_students"))

@app.route('/class-teacher')
def class_teacher_portal():
    db = get_db()
    cursor = db.cursor()
    current_session = get_current_session()
    current_term = get_current_term()
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Get all class arms for the dropdown
    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm, c.level
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """)
    classes = cursor.fetchall()
    
    return render_template('class_teacher_portal.html', 
                         classes=classes,
                         current_session=current_session,
                         current_term=current_term,
                         current_date=current_date)

def get_upload_status(class_arm_id, term, session, report_type='full_term'):
    """Get upload status for all subjects in a class"""
    db = get_db()
    cursor = db.cursor()

    # Get all subjects required for this class
    cursor.execute("""
        SELECT s.id, s.name, s.is_common_core, csr.is_compulsory
        FROM subjects s
        JOIN class_subject_requirements csr ON s.id = csr.subject_id
        WHERE csr.class_arm_id = ?
        ORDER BY s.name
    """, (class_arm_id,))
    subjects = cursor.fetchall()

    # Get number of students in the class
    cursor.execute("""
        SELECT COUNT(*) as student_count
        FROM student_classes
        WHERE class_arm_id = ? AND session = ? AND term = ?
    """, (class_arm_id, session, term))
    result = cursor.fetchone()
    total_students = result['student_count'] if result else 0

    upload_status = []

    for subject in subjects:
        # Count how many students have scores for this subject and report type
        cursor.execute("""
            SELECT COUNT(DISTINCT sc.student_id) as uploaded_count
            FROM scores sc
            JOIN student_classes sclass ON sc.student_id = sclass.student_id
            WHERE sclass.class_arm_id = ? 
              AND sc.subject_id = ? 
              AND sc.term = ? 
              AND sc.session = ? 
              AND sc.report_type = ?
        """, (class_arm_id, subject['id'], term, session, report_type))
        result = cursor.fetchone()
        uploaded_count = result['uploaded_count'] if result else 0

        # Calculate completion percentage
        completion_percentage = 0
        if total_students > 0:
            completion_percentage = (uploaded_count / total_students) * 100

        # Determine status
        if total_students == 0:
            status = 'no_students'
        elif uploaded_count == total_students:
            status = 'complete'
        elif uploaded_count > 0:
            status = 'partial'
        else:
            status = 'pending'

        upload_status.append({
            'subject_id': subject['id'],
            'subject_name': subject['name'],
            'is_required': subject['is_compulsory'],
            'uploaded_count': uploaded_count,
            'total_students': total_students,
            'completion_percentage': completion_percentage,
            'status': status
        })

    return upload_status

@app.route('/subject-teacher')
def subject_teacher_portal():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM subjects ORDER BY name")
    subjects = cursor.fetchall()

    # cursor.execute("""
    #     SELECT a.id AS arm_id, c.name AS class_name, a.arm
    #     FROM class_arms a
    #     JOIN classes c ON a.class_id = c.id
    #     ORDER BY c.id, a.arm
    # """)
    # classes = cursor.fetchall()

    classes_with_data = get_class_status()

    # Optional: fetch last 10 uploads
    cursor.execute("""
        SELECT s.name AS subject_name,
            c.name AS class_name,
            a.arm AS arm,
            sc.term,
            sc.session,
            MAX(sc.created_at) as date_uploaded
        FROM scores sc
        JOIN subjects s ON sc.subject_id = s.id
        JOIN class_arms a ON sc.class_arm_id = a.id
        JOIN classes c ON a.class_id = c.id
        GROUP BY s.name, c.name, a.arm, sc.term, sc.session
        ORDER BY date_uploaded DESC
        LIMIT 10
    """)
    recent_uploads = cursor.fetchall()

    current_session = get_current_session()

    return render_template(
        'subject_teacher_portal.html',
        subjects=subjects,
        classes=classes_with_data,
        current_session=current_session,
        recent_uploads=recent_uploads
    )

@app.route('/class-teacher/class/<int:class_arm_id>/<path:session>/<int:term>')
def class_teacher_class_view(class_arm_id, session, term):
    db = get_db()
    cursor = db.cursor()
    
    # Get class information
    cursor.execute("""
        SELECT c.name AS class_name, a.arm, c.level
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        WHERE a.id = ?
    """, (class_arm_id,))
    class_info = cursor.fetchone()
    
    # Get students in this class
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.age, s.gender, s.photo, s.department_id,
               d.name as department_name
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        LEFT JOIN departments d ON s.department_id = d.id
        WHERE sc.class_arm_id = ? AND sc.session = ?
        ORDER BY s.full_name
    """, (class_arm_id, session))
    students = cursor.fetchall()
    
    # Get attendance summary for each student
    # attendance_data = {}
    # for student in students:
    #     cursor.execute("""
    #         SELECT 
    #             COUNT(CASE WHEN status = 'present' THEN 1 END) as present,
    #             COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent,
    #             COUNT(CASE WHEN status = 'late' THEN 1 END) as late
    #         FROM attendance 
    #         WHERE student_id = ? AND class_arm_id = ? AND term = ? AND year = ?
    #     """, (student['id'], class_arm_id, term, year))
    #     attendance_data[student['id']] = cursor.fetchone()
    
    # Get assessment data
    assessment_data = {}
    for student in students:
        cursor.execute("""
            SELECT * FROM student_assessments 
            WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
        """, (student['id'], class_arm_id, term, session))
        assessment_data[student['id']] = cursor.fetchone()

    half_term_status = get_upload_status(class_arm_id, term, session, 'half_term')
    full_term_status = get_upload_status(class_arm_id, term, session, 'full_term')

    current_date = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('class_teacher_class_view.html',
                         class_info=class_info,
                         students=students,
                        #  attendance_data=attendance_data,
                        half_term_status=half_term_status,
                         full_term_status=full_term_status,
                         assessment_data=assessment_data,
                         class_arm_id=class_arm_id,
                         session=session,
                         term=term,
                         current_date=current_date) 

@app.route('/attendance/sheet/<int:class_arm_id>')
@app.route('/attendance/sheet/<int:class_arm_id>/<string:date>/<int:term>')
def attendance_sheet(class_arm_id, date=None, term=get_current_term()):
    db = get_db()
    cursor = db.cursor()
    
    # If date is not provided, use today's date
    if not date:
        date = datetime.now().strftime('%Y-%m-%d')
    
    # Get class information
    cursor.execute("""
        SELECT c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        WHERE a.id = ?
    """, (class_arm_id,))
    class_info = cursor.fetchone()
    
    if not class_info:
        return redirect(url_for('class_teacher_portal'))
    
    # Get students in class
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.photo
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.year = ?
        ORDER BY s.full_name
    """, (class_arm_id, datetime.now().year))
    students = cursor.fetchall()
    
    # Get existing attendance for this date
    cursor.execute("""
        SELECT student_id, status FROM attendance 
        WHERE class_arm_id = ? AND date = ? AND term = ? AND year = ?
    """, (class_arm_id, date, term, datetime.now().year))
    existing_attendance = {row['student_id']: row['status'] for row in cursor.fetchall()}
    
    return render_template('attendance_sheet.html',
                         students=students,
                         class_info=class_info,
                         class_arm_id=class_arm_id,
                         date=date,
                         term=term,
                         existing_attendance=existing_attendance)

@app.route('/attendance/submit', methods=['POST'])
def submit_attendance():
    class_arm_id = request.form['class_arm_id']
    date = request.form['date']
    term = request.form['term']
    year = datetime.now().year
    
    db = get_db()
    cursor = db.cursor()
    
    for key, value in request.form.items():
        if key.startswith('status_'):
            student_id = key.replace('status_', '')
            
            # Delete existing record for this student/date
            cursor.execute("""
                DELETE FROM attendance 
                WHERE student_id = ? AND class_arm_id = ? AND date = ? AND term = ? AND year = ?
            """, (student_id, class_arm_id, date, term, year))
            
            # Insert new record
            if value in ['present', 'absent', 'late']:
                cursor.execute("""
                    INSERT INTO attendance (student_id, class_arm_id, date, status, term, year)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (student_id, class_arm_id, date, value, term, year))
    
    db.commit()
    return redirect(url_for('class_teacher_portal'))

@app.route('/attendance/redirect')
def attendance_sheet_redirect():
    class_arm_id = request.args.get('class_arm_id')
    date = request.args.get('date')
    term = request.args.get('term')
    
    if not all([class_arm_id, date, term]):
        return redirect(url_for('class_teacher_portal'))
    
    return redirect(url_for('attendance_sheet', 
                        class_arm_id=class_arm_id, 
                        date=date, 
                        term=term))

@app.route('/assessments/bulk/<int:class_arm_id>/<int:term>/<path:session>')
def bulk_assessments(class_arm_id, term, session):
    db = get_db()
    cursor = db.cursor()
    
    # Get class information
    cursor.execute("""
        SELECT c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        WHERE a.id = ?
    """, (class_arm_id,))
    class_info = cursor.fetchone()
    
    # Get students in this class
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.photo
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ?
        ORDER BY s.full_name
    """, (class_arm_id, session))
    students = cursor.fetchall()
    
    # ---- get student averages -----------------------------------------
    cursor.execute("""
        SELECT 
            s.id AS student_id,
            ROUND(AVG(sc.total_score), 2) AS average_score
        FROM students s
        JOIN scores sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ?
        AND sc.term = ?
        AND sc.session = ?
        AND sc.report_type = 'full_term'
        GROUP BY s.id
    """, (class_arm_id, term, session))

    averages = {
        row['student_id']: row['average_score']
        for row in cursor.fetchall()
    }

    # Get existing assessment data
    assessment_data = {}
    for student in students:
        cursor.execute("""
            SELECT * FROM student_assessments 
            WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
        """, (student['id'], class_arm_id, term, session))
        result = cursor.fetchone()
        if result:
            assessment_data[student['id']] = dict(result)

    cursor.execute("""
        SELECT * FROM principal_comments
        ORDER BY min_average
    """)
    principal_comments = cursor.fetchall()

    cursor.execute("SELECT * FROM skills ORDER BY name")
    skills = cursor.fetchall()

    cursor.execute("""
        SELECT * FROM student_skills
        WHERE class_arm_id=? AND term=? AND session=?
    """, (class_arm_id, term, session))

    skill_data = {}
    for row in cursor.fetchall():
        skill_data.setdefault(row["student_id"], {})[row["skill_id"]] = row["score"]

    return render_template('bulk_assessments.j2',
                         class_info=class_info,
                         students=students,
                         averages=averages,
                         assessment_data=assessment_data,
                         class_arm_id=class_arm_id,
                         principal_comments=principal_comments,
                         skills=skills,
                         skill_data=skill_data,
                         term=term,
                         current_session=session)

@app.route('/assessments/submit-bulk', methods=['POST'])
def submit_bulk_assessments():
    class_arm_id = request.form['class_arm_id']
    term = request.form['term']
    session = get_current_session()
    
    db = get_db()
    cursor = db.cursor()
    
    # Get all students in this class
    cursor.execute("""
        SELECT s.id 
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ?
    """, (class_arm_id, session))
    students = cursor.fetchall()

    for student in students:
        student_id = student['id']
        
        # Extract assessment data for this student
        handwriting = request.form.get(f'handwriting_{student_id}')
        sports_participation = request.form.get(f'sports_participation_{student_id}')
        practical_skills = request.form.get(f'practical_skills_{student_id}')
        punctuality = request.form.get(f'punctuality_{student_id}')
        politeness = request.form.get(f'politeness_{student_id}')
        neatness = request.form.get(f'neatness_{student_id}')
        class_teacher_comment = request.form.get(f'class_teacher_comment_{student_id}', '')
        principal_comment = request.form.get(f'principal_comment_{student_id}', '')

        # Convert to integers if they exist
        handwriting = int(handwriting) if handwriting else None
        sports_participation = int(sports_participation) if sports_participation else None
        practical_skills = int(practical_skills) if practical_skills else None
        punctuality = int(punctuality) if punctuality else None
        politeness = int(politeness) if politeness else None
        neatness = int(neatness) if neatness else None
        
        # Insert or update assessment record
        cursor.execute('''INSERT OR REPLACE INTO student_assessments 
                          (student_id, class_arm_id, term, session, 
                           handwriting, sports_participation, practical_skills,
                           punctuality, politeness, neatness, class_teacher_comment, principal_comment) 
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (student_id, class_arm_id, term, session,
                        handwriting, sports_participation, practical_skills,
                        punctuality, politeness, neatness, class_teacher_comment, principal_comment))

        skill_id = request.form.get(f"skill_id_{student['id']}")
        raw_score = request.form.get(f"skill_score_{student['id']}")

        try:
            skill_id = int(skill_id)
            score = int(raw_score) if raw_score not in (None, "") else 0
        except ValueError:
            continue

        cursor.execute("""
            INSERT OR REPLACE INTO student_skills
            (student_id, class_arm_id, term, session, skill_id, score)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            student["id"],
            class_arm_id,
            term,
            session,
            skill_id,
            score
        ))

    db.commit()
    
    return redirect(url_for('class_teacher_class_view', 
                          class_arm_id=class_arm_id, 
                          session=session, 
                          term=term))

@app.route('/attendance/summary/<int:class_arm_id>/<int:term>/<path:session>')
def attendance_summary(class_arm_id, term, session):
    db = get_db()
    cursor = db.cursor()
    
    # Get class information
    cursor.execute("""
        SELECT c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        WHERE a.id = ?
    """, (class_arm_id,))
    class_info = cursor.fetchone()
    
    # Get students in this class
    cursor.execute("""
        SELECT s.id, s.reg_number, s.full_name, s.photo
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ?
        ORDER BY s.full_name
    """, (class_arm_id, session))
    students = cursor.fetchall()
    
    # Get existing attendance data
    attendance_data = {}
    total_school_days = None
    
    for student in students:
        cursor.execute("""
            SELECT days_present, days_absent, days_late, total_school_days
            FROM attendance_summary 
            WHERE student_id = ? AND class_arm_id = ? AND term = ? AND session = ?
        """, (student['id'], class_arm_id, term, session))
        result = cursor.fetchone()
        if result:
            attendance_data[student['id']] = dict(result)
            # Use the total_school_days from the first student record
            if total_school_days is None:
                total_school_days = result['total_school_days']
    
    return render_template('attendance_summary.j2',
                         class_info=class_info,
                         students=students,
                         attendance_data=attendance_data,
                         total_school_days=total_school_days,
                         class_arm_id=class_arm_id,
                         term=term,
                         session=session)

@app.route('/attendance/submit-summary', methods=['POST'])
def submit_attendance_summary():
    class_arm_id = request.form['class_arm_id']
    term = request.form['term']
    session = get_current_session()
    total_school_days = int(request.form['total_school_days'])
    
    db = get_db()
    cursor = db.cursor()
    
    # Get all students in this class
    cursor.execute("""
        SELECT s.id 
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ?
    """, (class_arm_id, session))
    students = cursor.fetchall()
    
    for student in students:
        student_id = student['id']
        
        # Extract attendance data for this student
        days_present = int(request.form.get(f'present_{student_id}', 0))
        days_absent = int(request.form.get(f'absent_{student_id}', 0))
        days_late = int(request.form.get(f'late_{student_id}', 0))
        
        # Insert or update attendance summary
        cursor.execute('''INSERT OR REPLACE INTO attendance_summary 
                          (student_id, class_arm_id, term, session, 
                           days_present, days_absent, days_late, total_school_days) 
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                       (student_id, class_arm_id, term, session,
                        days_present, days_absent, days_late, total_school_days))
    
    db.commit()
    
    return redirect(url_for('class_teacher_class_view', 
                          class_arm_id=class_arm_id, 
                          session=session, 
                          term=term))


@app.route('/student-photos')
def student_photos():
    db = get_db()
    cursor = db.cursor()
    
    # Get all classes for dropdown
    cursor.execute("""
        SELECT a.id AS arm_id, c.name AS class_name, a.arm
        FROM class_arms a
        JOIN classes c ON a.class_id = c.id
        ORDER BY c.id, a.arm
    """)
    classes = cursor.fetchall()
    
    selected_class = request.args.get('class_arm_id')
    students = []
    session = get_current_session()

    if selected_class:
        # Get students in selected class for current year
        cursor.execute("""
            SELECT s.reg_number, s.full_name, s.photo, 
                   c.name AS class_name, a.arm
            FROM students s
            JOIN student_classes sc ON s.id = sc.student_id
            JOIN class_arms a ON sc.class_arm_id = a.id
            JOIN classes c ON a.class_id = c.id
            WHERE sc.class_arm_id = ? AND sc.session = ?
            ORDER BY s.full_name
        """, (selected_class, session))
        students = cursor.fetchall()
    
    return render_template('photo_management.html',
                         classes=classes,
                         students=students,
                         selected_class=selected_class)

#Testing
@app.route('/generate-test-results')
def generate_test_results():
    """
    Generate test full-term result Excel files for all subjects in a class.
    Uses real student names from the DB.
    """
    class_arm_id = request.args.get("class_arm_id", type=int)
    session = request.args.get("session")
    term = request.args.get("term", type=int)

    if not all([class_arm_id, session, term]):
        return "Missing parameters: class_arm_id, session, term", 400

    db = get_db()
    cursor = db.cursor()

    # Get students in this class
    cursor.execute("""
        SELECT s.full_name
        FROM students s
        JOIN student_classes sc ON s.id = sc.student_id
        WHERE sc.class_arm_id = ? AND sc.session = ? AND sc.term = ?
        ORDER BY s.full_name
    """, (class_arm_id, session, term))
    students = cursor.fetchall()

    if not students:
        return "No students found for that class/session/term", 404

    # Get subjects for that class
    cursor.execute("""
        SELECT s.id, s.name
        FROM subjects s
        JOIN class_subject_requirements csr ON csr.subject_id = s.id
        WHERE csr.class_arm_id = ?
        ORDER BY s.name
    """, (class_arm_id,))
    subjects = cursor.fetchall()

    if not subjects:
        return "No subjects found for this class", 404

    # Prepare ZIP file in memory
    zip_buffer = BytesIO()
    import zipfile

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for subject in subjects:
            subject_id = subject["id"]
            subject_name = subject["name"].replace(" ", "_")

            # Generate random scores
            data = {
                "full_name": [s["full_name"] for s in students],
                "ca1": np.random.randint(0, 6, len(students)),
                "ca2": np.random.randint(0, 6, len(students)),
                "ca3": np.random.randint(0, 6, len(students)),
                "ca4": np.random.randint(0, 6, len(students)),
                "exam": np.random.randint(40, 81, len(students)),
            }

            df = pd.DataFrame(data)
            df["total"] = df["ca1"] + df["ca2"] + df["ca3"] + df["ca4"] + df["exam"]

            # Write to Excel (in memory)
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Results", index=False)
            output.seek(0)

            # Add Excel to ZIP
            filename = f"{subject_name}_full_term_test_results.xlsx"
            zf.writestr(filename, output.getvalue())

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"class_{class_arm_id}_full_term_test_results.zip"
    )

def allowed_photo_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'jpg', 'jpeg', 'png', 'gif'}

with app.app_context():
    init_db()

if __name__ == "__main__":
    init_db()
    # Automatically get your local network IP address
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)

    print(f"\n🚀 Server running on:")
    print(f"   Local:     http://127.0.0.1:5000")
    print(f"   Network:   http://{local_ip}:5000\n")
    print("📡 Connect other devices on the same Wi-Fi using the 'Network' address above.\n")

    # Run Flask on all interfaces (so others can access it)
    app.run(host="0.0.0.0", port=5000, debug=True)
